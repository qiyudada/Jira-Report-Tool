"""
Report Generator - Issue filtering and Excel generation
"""
import datetime
import re
import os
import json
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import timedelta

from .blocked import compute_blocked_marker
from .config import Config


class ReportGenerator:
    ENGINEER_FIELD = "Software Development Engineer 软件开发工程师"

    def __init__(self, jira_client, config):
        self.client = jira_client
        self.config = config
        self.base_url = jira_client.base_url

    def generate(self, start_date: datetime.date, end_date: datetime.date,
                 output_path: str, status_filter: str = "ALL",
                 use_ai_summary: bool = False,
                 fetch_comment: bool = False,
                 batch_mode: bool = False,
                 batch_size: int = 10,
                 ai_summarizer=None):
        """Generate Jira report Excel file (legacy one-shot path used by
        `cli.py run --ai` and direct Python callers). For the skill/agent flow,
        use `collect_issues_data` + `export_from_data` instead."""
        issues = self._fetch_and_filter_issues(start_date, end_date, status_filter)

        # Create Excel
        self._create_excel(issues, output_path, status_filter, start_date, end_date,
                          use_ai_summary, fetch_comment, batch_mode, batch_size, ai_summarizer)

        return len(issues)

    def _fetch_and_filter_issues(self, start_date: datetime.date, end_date: datetime.date,
                                  status_filter: str) -> list:
        """Fetch issues across all JQL queries, filter, deduplicate, and sort.
        Shared by `generate` (legacy) and `collect_issues_data` (skill flow)."""
        status_clause = f'status = "{status_filter}" ' if status_filter != "ALL" else ""

        jql_normal = f'"{self.ENGINEER_FIELD}" IN (currentUser()) AND updated >= {start_date} AND updated <= "{end_date} 23:59"'
        if status_clause:
            jql_normal += f' AND {status_clause}'

        jql_wait3rd = f'"{self.ENGINEER_FIELD}" IN (currentUser()) AND status = "WAIT 3RD PARTY" AND updated >= {start_date} AND updated <= "{end_date} 23:59"'
        if status_clause:
            jql_wait3rd += f' AND {status_clause}'

        jql_assist_normal = f'comment ~ currentUser() AND "{self.ENGINEER_FIELD}" != currentUser() AND updated >= {start_date} AND updated <= "{end_date} 23:59"'
        if status_clause:
            jql_assist_normal += f' AND {status_clause}'

        jql_assist_wait3rd = f'comment ~ currentUser() AND "{self.ENGINEER_FIELD}" != currentUser() AND status = "WAIT 3RD PARTY" AND updated >= {start_date} AND updated <= "{end_date} 23:59"'
        if status_clause:
            jql_assist_wait3rd += f' AND {status_clause}'

        st_bug_review_values = [
            "ST下一版本验证", "ST最新版本验证", "不再修复",
            "升A修复", "升V修复", "暂不修复",
        ]
        st_bug_review_clause = ", ".join(f'"{value}"' for value in st_bug_review_values)
        jql_st_bug_review = (
            f'"ST BUG评估意见" in ({st_bug_review_clause}) '
            f'AND updated >= {start_date} AND updated <= "{end_date} 23:59" '
            f'AND assignee in (currentUser()) ORDER BY updated DESC'
        )

        # New: assignee-based query — tickets assigned to current user regardless of SDE field.
        # Post-fetch filter requires the current user to have commented in the date range.
        jql_assignee = f'assignee in (currentUser()) AND updated >= {start_date} AND updated <= "{end_date} 23:59"'
        if status_clause:
            jql_assignee += f' AND {status_clause}'

        issues_assigned_normal = self.client.fetch_issues(jql_normal)
        issues_assigned_wait3rd = self.client.fetch_issues(jql_wait3rd)
        issues_assigned = issues_assigned_normal + issues_assigned_wait3rd

        issues_assist_normal = self.client.fetch_issues(jql_assist_normal)
        issues_assist_wait3rd = self.client.fetch_issues(jql_assist_wait3rd)
        issues_assist = issues_assist_normal + issues_assist_wait3rd

        issues_st_bug_review = self.client.fetch_issues(jql_st_bug_review)

        issues_assigned_filtered = [
            issue for issue in issues_assigned
            if self._should_include_issue(issue, start_date, end_date)
        ]
        issues_assist_filtered = [
            issue for issue in issues_assist
            if self._should_include_issue(issue, start_date, end_date)
        ]

        # Fetch assignee-based issues and keep only those where user commented in range.
        issues_assignee = self.client.fetch_issues(jql_assignee)
        issues_assignee_filtered = [
            issue for issue in issues_assignee
            if self.client.user_commented_in_date_range(issue['key'], start_date, end_date)
        ]

        # Build deduplicated dict: existing queries take priority (added first),
        # assignee-based issues fill in any remaining gaps.
        all_issues = {issue['key']: issue for issue in (
            issues_assigned_filtered + issues_assist_filtered + issues_st_bug_review + issues_assignee_filtered
        )}
        issues = list(all_issues.values())
        issues.sort(key=lambda x: -self._get_created_timestamp(x))
        return issues

    def collect_issues_data(self, start_date: datetime.date, end_date: datetime.date,
                             status_filter: str = "ALL") -> dict:
        """Fetch + filter + collect comments. Emit a JSON-serializable dict the
        agent (or test code) can consume. No AI is called here — `prefilled_summary`
        is the only automatic 进展 text, applied to blocked-status issues with no
        in-period activity."""
        issues = self._fetch_and_filter_issues(start_date, end_date, status_filter)
        context_start = end_date - timedelta(days=60)

        out_issues = []
        for issue in issues:
            issue_key = issue.get("key", "")
            fields = issue.get("fields", {})
            status = (fields.get("status") or {}).get("name", "")
            comments = self._get_all_comments_in_range(
                issue_key, start_date, end_date, context_start=context_start,
            )
            customer_name, model_name = self._resolve_customer_and_model(issue_key, fields)
            key_issue_value, highlight_key_issue = self._resolve_key_issue_value(issue_key, fields)
            prefilled = compute_blocked_marker(status, comments, end_date) or ""

            out_issues.append({
                "key": issue_key,
                "summary": fields.get("summary", ""),
                "status": status,
                "customer_name": customer_name,
                "model_name": model_name,
                "key_issue": key_issue_value,
                "highlight_key_issue": highlight_key_issue,
                "comments": [
                    {
                        "author": c.get("author", ""),
                        "date": c["date"].isoformat() if hasattr(c.get("date"), "isoformat") else c.get("date"),
                        "body": c.get("body", ""),
                        "in_period": c.get("in_period", True),
                        "author_role": c.get("author_role", ""),
                    }
                    for c in comments
                ],
                "prefilled_summary": prefilled,
            })

        return {
            "metadata": {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
                "status_filter": status_filter,
                "column_order": self.config.column_order,
                "key_issue_highlight": self.config.key_issue_highlight,
                "header_align": self.config.header_align,
                "cell_align": self.config.cell_align,
                "jira_base_url": Config.JIRA_BASE_URL,
            },
            "issues": out_issues,
        }

    def export_from_data(self, data: dict, summaries: dict, output_path: str) -> int:
        """Render an Excel report from a previously-collected `data` dict and an
        agent-supplied `summaries` mapping (issue_key -> summary text).
        `prefilled_summary` (if set) takes precedence over `summaries`. Missing
        or empty entries in `summaries` produce an empty 进展 cell (with a stderr
        warning). Returns the number of issues written."""
        metadata = data.get("metadata") or {}
        start_date = datetime.date.fromisoformat(metadata["start_date"])
        end_date = datetime.date.fromisoformat(metadata["end_date"])
        status_filter = metadata.get("status_filter", "ALL")

        out_issues = []
        latest_comments = {}
        expected_keys = set()
        for issue in data.get("issues", []):
            issue_key = issue["key"]
            expected_keys.add(issue_key)
            prefilled = (issue.get("prefilled_summary") or "").strip()
            if prefilled:
                latest_comments[issue_key] = prefilled
            elif issue_key in summaries:
                text = (summaries[issue_key] or "").strip()
                if text:
                    latest_comments[issue_key] = text
                else:
                    print(f"Warning: empty summary for {issue_key}, leaving 进展 blank.", file=sys.stderr)
            else:
                print(f"Warning: missing summary for {issue_key}, leaving 进展 blank.", file=sys.stderr)
            out_issues.append(self._issue_dict_to_jira_shape(issue))

        extra_keys = set(summaries.keys()) - expected_keys
        for k in extra_keys:
            print(f"Warning: summary provided for unknown issue {k}, ignoring.", file=sys.stderr)

        # Apply metadata-derived config overrides so the rendered Excel matches
        # the prepare-time settings, regardless of the current Config object.
        self._apply_metadata_overrides(metadata)

        self._create_excel(
            out_issues, output_path, status_filter, start_date, end_date,
            use_ai_summary=False, fetch_comment=False,
            precomputed_latest_comments=latest_comments,
        )
        return len(out_issues)

    def _issue_dict_to_jira_shape(self, issue: dict) -> dict:
        """Reconstruct the {'key', 'fields': {...}} shape that _create_excel
        expects from a flat issue dict in the JSON data file."""
        fields = {
            "summary": issue.get("summary", ""),
            "status": {"name": issue.get("status", "")},
        }
        # Re-supply the original field ids that the resolver helpers read.
        # _resolve_customer_and_model falls back to "" when these are missing,
        # so the prepare-supplied customer_name/model_name survive.
        fields["customfield_11029"] = issue.get("customer_name", "")
        fields["customfield_12031"] = issue.get("model_name", "")
        fields["customfield_10102"] = issue.get("customer_name", "")
        fields["customfield_10400"] = issue.get("model_name", "")
        fields["customfield_10401"] = issue.get("model_name", "")
        fields["customfield_11043"] = issue.get("model_name", "")
        fields["customfield_11044"] = issue.get("key_issue", "")
        fields["priority"] = {"name": "High" if issue.get("highlight_key_issue") else "Medium"}
        fields["issuetype"] = {"name": ""}
        return {"key": issue["key"], "fields": fields}

    def _apply_metadata_overrides(self, metadata: dict):
        """Temporarily override config fields from `metadata` so the rendered
        Excel matches the prepare-time settings. Touches only string fields;
        persistent config on `self.config` is mutated in place but the caller
        is expected to be a one-shot CLI invocation."""
        if "column_order" in metadata:
            self.config.column_order = metadata["column_order"]
        if "key_issue_highlight" in metadata:
            self.config.key_issue_highlight = bool(metadata["key_issue_highlight"])
        if "header_align" in metadata:
            self.config.header_align = metadata["header_align"]
        if "cell_align" in metadata:
            self.config.cell_align = metadata["cell_align"]

    def _get_created_timestamp(self, issue):
        created_str = issue.get("fields", {}).get("created", "")
        dt = self.client._parse_jira_datetime(created_str)
        return dt.timestamp() if dt else 0

    def _should_include_issue(self, issue, start_date, end_date):
        """Determine if an issue should be included in the report"""
        status = issue.get("fields", {}).get("status", {}).get("name", "")
        status_key = status.upper()
        fields = issue.get("fields", {})
        assignee_field = fields.get("assignee")
        sde_field = fields.get("customfield_12001")

        no_comment_required_statuses = {"WAIT 3RD PARTY", "WORKING"}
        wait_blocked_statuses = {"WAIT FAE INFO", "WORKED AROUND", "WAIT OFFICIAL RELEASE"}
        closed_statuses = {"CLOSED", "RESOLVED"}

        if status_key in no_comment_required_statuses:
            return self._is_in_date_range(issue, start_date, end_date)

        if status_key in wait_blocked_statuses:
            if self.client.is_current_user(sde_field):
                if not self.client.is_current_user(assignee_field):
                    return True
                return self.client.user_commented_in_date_range(issue['key'], start_date, end_date)
            return self.client.user_commented_in_date_range(issue['key'], start_date, end_date)

        if status_key in closed_statuses and self.client.is_current_user(sde_field):
            if self._field_date_in_range(issue, "resolutiondate", start_date, end_date) or \
               self._field_date_in_range(issue, "updated", start_date, end_date):
                return True

        return self.client.user_commented_in_date_range(issue['key'], start_date, end_date)

    def _is_in_date_range(self, issue, start_date, end_date):
        created_str = issue.get("fields", {}).get("created", "")
        dt = self.client._parse_jira_datetime(created_str)
        if dt and start_date <= dt.date() <= end_date:
            return True
        return self.client.user_commented_in_date_range(issue['key'], start_date, end_date)

    def _field_date_in_range(self, issue, field_name, start_date, end_date):
        date_str = issue.get("fields", {}).get(field_name, "")
        dt = self.client._parse_jira_datetime(date_str)
        return bool(dt and start_date <= dt.date() <= end_date)

    def _is_jira_issue_key_text(self, text):
        return bool(re.fullmatch(r'[A-Z][A-Z0-9]+-\d+', str(text or "").strip(), re.IGNORECASE))

    def _is_st_issue(self, issue_key):
        return str(issue_key or "").upper().startswith("ST")

    def _resolve_key_issue_value(self, issue_key, fields):
        """Resolve the Excel key-issue column value and whether it should be highlighted"""
        if self._is_st_issue(issue_key):
            severity = self.client._field_to_text(fields.get("customfield_11044"))
            match = re.match(r'\s*([A-E])(?:\b|-)', severity, flags=re.IGNORECASE)
            severity_level = match.group(1).upper() if match else severity.strip()
            return severity_level, severity_level in ("A", "B")

        priority = fields.get("priority", {})
        priority_name = priority.get("name", "") if isinstance(priority, dict) else priority
        is_key_issue = priority_name in ("Highest", "High")
        return "是" if is_key_issue else "否", is_key_issue

    def _extract_epic_display_name(self, fields):
        """Extract readable epic title"""
        epic_name = self.client._field_to_text(fields.get("customfield_10102"))
        if epic_name and not self._is_jira_issue_key_text(epic_name):
            return epic_name

        epic_link = fields.get("customfield_10100")
        if isinstance(epic_link, dict):
            for key in ("summary", "name", "value"):
                text = str(epic_link.get(key, "") or "").strip()
                if text and not self._is_jira_issue_key_text(text):
                    return text

        text = self.client._field_to_text(epic_link)
        if text and not self._is_jira_issue_key_text(text):
            return text
        return ""

    def _resolve_customer_and_model(self, issue_key, fields):
        """Resolve customer and model values for export"""
        issue_key = str(issue_key or "")
        is_st_issue = self._is_st_issue(issue_key)
        is_rd_issue = issue_key.upper().startswith("SW")
        is_opendemand = issue_key.upper().startswith("OPENDEMAND")

        customer = self.client._field_to_text(fields.get("customfield_11029"))
        model = self.client._field_to_text(fields.get("customfield_12031"))

        issue_type_field = fields.get("issuetype")
        if isinstance(issue_type_field, dict):
            issue_type = str(issue_type_field.get("name", "") or "").strip()
        else:
            issue_type = self.client._field_to_text(issue_type_field)
        st_model = self.client._field_to_text(fields.get("customfield_11043"))
        epic_name = self._extract_epic_display_name(fields)
        platform = self.client._field_to_text(fields.get("customfield_10400")) or \
                   self.client._field_to_text(fields.get("customfield_10401"))

        if is_opendemand:
            # OPENDEMAND (需求评估) tickets: use issue type as customer, model left blank.
            customer = issue_type or customer
            model = ""
        elif is_st_issue:
            customer = issue_type or customer
            model = st_model or model
        elif is_rd_issue:
            customer = epic_name or customer
            model = platform or model
        else:
            if not customer:
                customer = epic_name
            if not model:
                model = platform

        return customer, model

    def _create_excel(self, issues, filepath, statuses, start_date, end_date,
                     use_ai_summary=False, fetch_comment=False,
                     batch_mode=False, batch_size=10, ai_summarizer=None,
                     precomputed_latest_comments=None):
        """Create Excel report file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        font_chinese = Font(name="Microsoft YaHei", size=10)
        font_english = Font(name="JetBrains Mono", size=10)
        font_header = Font(name="Microsoft YaHei", bold=True, color="FF000000", size=10)

        header_align = self.config.header_align
        cell_align = self.config.cell_align
        header_alignment = Alignment(horizontal=header_align, vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal=cell_align, vertical="center", wrap_text=True)

        thin = Side(border_style="thin", color="FF000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        col_order = [int(x.strip()) - 1 for x in self.config.column_order.split(",")]

        header_names = ["客户名称", "型号", "问题描述",
                        "JIRA号", "状态", "是否为重点问题",
                        "进展"]

        def has_chinese(text):
            return any('一' <= c <= '鿿' for c in str(text))

        def set_cell_font(cell, value):
            text = str(value) if value is not None else ""
            cell.value = text if text else value
            cell.font = font_chinese if has_chinese(text) else font_english

        for col, idx in enumerate(col_order, 1):
            header = header_names[idx]
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = font_header
            cell.alignment = header_alignment
            cell.border = border

        status_col = col_order.index(4) + 1
        key_issue_col = col_order.index(5) + 1
        status_range = f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{len(issues) + 1}"
        key_issue_range = f"{get_column_letter(key_issue_col)}2:{get_column_letter(key_issue_col)}{len(issues) + 1}"

        ws_options = wb.create_sheet("_Options")
        ws_options.sheet_state = "hidden"
        for i, opt in enumerate(["WAIT FAE INFO", "WORKED AROUND", "WORKING", "CLOSED", "RESOLVED", "WAIT 3RD PARTY"], 1):
            ws_options.cell(row=i, column=1, value=opt)
        for i, opt in enumerate(["是", "否", "A", "B", "C", "D", "E"], 1):
            ws_options.cell(row=i, column=2, value=opt)
        status_options_range = "_Options!$A$1:$A$6"
        key_issue_options_range = "_Options!$B$1:$B$7"

        dv_status = DataValidation(type="list", formula1=status_options_range, allow_blank=True)
        dv_status.error = "Please select a valid status"
        dv_status.errorTitle = "Invalid Status"
        ws.add_data_validation(dv_status)
        dv_status.sqref = status_range

        dv_key_issue = DataValidation(type="list", formula1=key_issue_options_range, allow_blank=True)
        dv_key_issue.error = "Please select 是 or 否"
        dv_key_issue.errorTitle = "Invalid Key Issue"
        ws.add_data_validation(dv_key_issue)
        dv_key_issue.sqref = key_issue_range

        # Pre-fetch AI summaries if enabled. `precomputed_latest_comments` is
        # the agent-supplied path (skill flow): when given, skip the AI entirely
        # and trust the caller-provided mapping.
        latest_comments = precomputed_latest_comments if precomputed_latest_comments is not None else {}
        if precomputed_latest_comments is None and use_ai_summary and ai_summarizer:
            context_start = end_date - timedelta(days=60)
            issues_data = []
            for issue in issues:
                issue_key = issue.get("key", "")
                fields = issue.get("fields", {})
                all_comments = self._get_all_comments_in_range(
                    issue_key, start_date, end_date, context_start=context_start
                )
                issues_data.append({
                    "issue_key": issue_key,
                    "summary": fields.get("summary", ""),
                    "comments": all_comments
                })

            if batch_mode:
                batch_size = max(1, batch_size)
                for i in range(0, len(issues_data), batch_size):
                    batch = issues_data[i:i+batch_size]
                    results = ai_summarizer.batch_summarize(batch, self.config.ai_model)
                    latest_comments.update(results)
            else:
                for item in issues_data:
                    ai_summary = ai_summarizer.summarize(
                        item['issue_key'],
                        item['summary'],
                        item['comments'],
                        self.config.ai_model
                    )
                    latest_comments[item['issue_key']] = ai_summary

        # Write data rows
        for row, issue in enumerate(issues, 2):
            fields = issue.get("fields", {})
            issue_key = issue.get("key", "")
            customer_name, model_name = self._resolve_customer_and_model(issue_key, fields)

            latest_comment = latest_comments.get(issue_key, "")
            if fetch_comment and not latest_comment:
                latest_comment = self.client.get_user_latest_comment(
                    issue_key, start_date, end_date,
                    self.config.comment_timestamp_prefix
                ) or ""

            key_issue_value, highlight_key_issue = self._resolve_key_issue_value(issue_key, fields)

            values = [
                customer_name,
                model_name,
                fields.get("summary", ""),
                issue_key,
                fields.get("status", {}),
                key_issue_value,
                latest_comment,
            ]

            for col, idx in enumerate(col_order, 1):
                val = values[idx]
                if idx == 4:
                    val = val.get("name", "") if isinstance(val, dict) else val

                cell = ws.cell(row=row, column=col, value=val)
                if idx == 3:
                    cell.hyperlink = f"{self.base_url}/browse/{issue_key}"
                set_cell_font(cell, val)
                cell.alignment = cell_alignment
                cell.border = border
                if idx == 5 and highlight_key_issue and self.config.key_issue_highlight:
                    cell.fill = PatternFill(fill_type="solid", start_color="FFFFC7CE", end_color="FFFFC7CE")

        # Auto-fit column widths
        for col in range(1, 8):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        max_length = max(max_length, cell_len)
                except:
                    pass
            adjusted_width = min(max_length + 5, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.row_dimensions[1].height = 25

        # Ensure output directory exists
        save_dir = os.path.dirname(filepath)
        if save_dir and not os.path.exists(save_dir):
            os.makedirs(save_dir)

        wb.save(filepath)

    def _get_all_comments_in_range(self, issue_key, start_date, end_date, context_start=None):
        """Get all comments within date range, optionally including background context"""
        try:
            comments = self.client.get_comments(issue_key)
            fetch_start = context_start if context_start else start_date

            result = []
            for comment in comments:
                author = comment.get("author", {})
                author_name = author.get("name") or author.get("displayName", "Unknown")
                author_email = author.get("emailAddress", "")
                created_str = comment.get("created", "")
                if not created_str:
                    continue
                comment_dt = self.client._parse_jira_datetime(created_str)
                if not comment_dt:
                    continue
                comment_date = comment_dt.date()
                if fetch_start <= comment_date <= end_date:
                    body = comment.get("body", "") or ""
                    text = self._clean_comment_body(body)
                    if text:
                        result.append({
                            "author": author_name,
                            "author_role": self._get_comment_author_role(author_name, author_email),
                            "date": comment_date,
                            "body": text,
                            "in_period": start_date <= comment_date <= end_date,
                        })
            return result
        except Exception:
            return []

    def _clean_comment_body(self, body):
        """Clean comment body: remove HTML, markup, attachments, emoji"""
        if not body:
            return ""

        body = re.sub(r'<div[^>]*>.*?</div>', '', body, flags=re.DOTALL)
        body = re.sub(r'<img[^>]*>', '', body)
        body = re.sub(r'<a[^>]*href=[^>]*>[^<]*</a>', '', body)
        body = re.sub(r'<[^>]+>', '', body)
        body = (body.replace('&nbsp;', ' ').replace('&amp;', '&')
                    .replace('&lt;', '<').replace('&gt;', '>')
                    .replace('&quot;', '"').replace('&#39;', "'"))

        body = re.sub(r'\{(?:panel|code|noformat|color|quote)[^}]*\}.*?\{(?:panel|code|noformat|color|quote)\}', '', body, flags=re.DOTALL | re.IGNORECASE)
        body = re.sub(r'\{[a-z][^}]{0,40}\}', '', body, flags=re.IGNORECASE)
        body = re.sub(r'\[~[^\]]+\]', '', body)
        body = re.sub(r'^h[1-6]\.\s*', '', body, flags=re.MULTILINE)
        body = re.sub(r'[*_+\-^~](\S[^*_+\-^~\n]*?\S)[*_+\-^~]', r'\1', body)

        body = re.sub(r'![\w\-\. ]+(?:\|[^!]*)!', '', body)
        body = re.sub(r'https?://\S+?\.(?:jpg|jpeg|png|gif|webp|bmp|svg)(?:\?\S*)?', '', body, flags=re.IGNORECASE)
        body = re.sub(r'\[\^[^\]]+\]', '', body)
        body = re.sub(r'_\(\d+[KMG]?B?\)_', '', body)
        body = re.sub(r'\[[^\]]*\.(?:zip|gz|rar|7z|tar|dmp|log|txt|pdf|doc|docx|xls|xlsx|ild)[^\]]*\]', '', body, flags=re.IGNORECASE)
        body = re.sub(r'(?:attachment|附件|文件)[：:\s][^\s,，\n]+', '', body, flags=re.IGNORECASE)

        body = re.compile(
            "[\U0001F600-\U0001F64F\U0001F300-\U0001F5FF\U0001F680-\U0001F6FF"
            "\U0001F700-\U0001F77F\U0001F780-\U0001F7FF\U0001F800-\U0001F8FF"
            "\U0001F900-\U0001F9FF\U0001FA00-\U0001FA6F\U0001FA70-\U0001FAFF"
            "☀-➿]+", flags=re.UNICODE
        ).sub('', body)

        _NOISE_PATTERNS = [
            r'(?:确认)?该issue(?:已)?由父issue\s*\S+\s*(?:通过复制\s*)?同步创建[^\n。]*',
            r'从父(?:issue|任务)\s*(?:\S+\s*)?(?:通过复制\s*)?同步创建[^\n。]*',
            r'由父issue[^\n。]*同步创建[^\n。]*',
            r'通过复制同步创建[^\n。]*',
            r'[Tt]his issue was created from parent issue\s*\S+\s*by a Copy\s*[&＆]?\s*Sync operation[^\n]*',
            r'[Cc]reated from parent issue[^\n]*(?:Copy|Sync)[^\n]*',
            r'[Ss]ynced? (?:from|with) parent issue[^\n]*',
            r'[Tt]his issue requires your attention[^\n]*',
            r'\b(?:Log|dbg|debug|dump|trace)[^。\n]*\.(?:zip|rar|7z|tar|gz|log|dmp|txt)\s*\([^)]*\)',
            r'报告了一个bug[^\n。]*',
            r'创建了(?:case|issue|工单)[^\n。]*',
            r'在Jira中进行了记录[^\n。]*',
            r'提供了日志用于分析[^\n。]*',
            r'(?:已)?进行了记录[^\n。]*',
            r'确认该issue[^\n。]*',
            r'请(?:帮忙)?关闭(?:该)?(?:case|issue|工单)[^\n。]*',
            r'已回复[^\n。]{0,20}',
            r'(?:Permalink|Edit|Delete|added a comment)[^\n]*',
        ]
        for pattern in _NOISE_PATTERNS:
            body = re.sub(pattern, '', body, flags=re.IGNORECASE)

        body = re.sub(r'\s+', ' ', body).strip()
        if body and not re.search(r'[\w一-鿿]', body):
            return ""

        return body

    def _get_comment_author_role(self, author_name, author_email):
        """Classify comment author role"""
        author_email = (author_email or "").lower()
        author_name = (author_name or "").lower()
        username = (self.client.username or "").lower()
        username_short = username.split("@")[0] if username else ""

        if username and (author_email == username or author_name == username or author_name == username_short):
            return "当前用户"
        if author_email.endswith("@quectel.com"):
            return "我方"
        return "客户/Reporter"