"""
Jira Weekly Report Generator
Desktop app to generate Excel reports from Jira issues
Modern AI Tool UI - MiniMax/OpenAI Style
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import requests
import requests.auth
import json
import datetime
from datetime import timedelta
import calendar
import os
import re
import threading
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# MiniMax-inspired dark product theme
THEME_BG = "#050505"
THEME_SURFACE = "#101010"
THEME_SURFACE_RAISED = "#181818"
THEME_SURFACE_HOVER = "#242424"
THEME_BORDER = "#2f2f2f"
THEME_PRIMARY = "#c7ff3d"
THEME_PRIMARY_HOVER = "#d8ff69"
THEME_PRIMARY_GLOW = "#8cff00"
THEME_PRIMARY_TEXT = "#070707"
THEME_TEXT = "#f4f4f4"
THEME_TEXT_SECONDARY = "#b7b7b7"
THEME_TEXT_MUTED = "#747474"
THEME_SUCCESS = "#4ec9b0"
THEME_ERROR = "#ff5c5c"
THEME_WARNING = "#f2c94c"
THEME_ACTIVITY_BAR = "#0b0b0b"
THEME_STATUS_BAR = "#101010"

# Backward compatibility
MINECRAFT_BG = THEME_BG
MINECRAFT_SURFACE = THEME_SURFACE
MINECRAFT_SURFACE_ALT = THEME_SURFACE_RAISED
MINECRAFT_BORDER = THEME_BORDER
MINECRAFT_GRASS = THEME_PRIMARY
MINECRAFT_GRASS_DARK = THEME_PRIMARY_HOVER
MINECRAFT_STONE = THEME_SURFACE
MINECRAFT_COBBLE = THEME_BORDER
MINECRAFT_LAVA = THEME_ERROR
MINECRAFT_WATER = THEME_PRIMARY
MINECRAFT_GOLD = THEME_WARNING
MINECRAFT_TEXT = THEME_TEXT
MINECRAFT_TEXT_DIM = THEME_TEXT_SECONDARY
MINECRAFT_GREEN = THEME_SUCCESS
MINECRAFT_RED = THEME_ERROR
MINECRAFT_YELLOW = THEME_WARNING
MINECRAFT_AQUA = THEME_PRIMARY
CHECKBOX_SELECT_BG = THEME_PRIMARY

VSCODE_BG = THEME_BG
VSCODE_SURFACE = THEME_SURFACE
VSCODE_SURFACE_ALT = THEME_SURFACE_RAISED
VSCODE_BORDER = THEME_BORDER
VSCODE_BLUE = THEME_PRIMARY
VSCODE_CYAN = THEME_PRIMARY
VSCODE_ORANGE = THEME_WARNING
VSCODE_GREEN = THEME_SUCCESS
VSCODE_RED = THEME_ERROR
VSCODE_YELLOW = THEME_WARNING
VSCODE_TEXT = THEME_TEXT
VSCODE_TEXT_DIM = THEME_TEXT_SECONDARY
VSCODE_DISABLED = THEME_TEXT_MUTED
VSCODE_SELECT = THEME_PRIMARY


class OperationCancelled(Exception):
    """Raised when the user cancels report generation."""


class JiraReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Jira Report")
        self.root.geometry("1040x680")
        self.root.minsize(920, 620)
        self.root.configure(bg=VSCODE_BG)

        # Jira API settings
        self.base_url = "https://ticket.quectel.com"
        self.session = requests.Session()
        self.logged_in = False
        self.username = None
        self.user_email = None
        self.cancel_event = threading.Event()
        self.generation_running = False

        # Config file
        self.config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".jira_config")
        self.last_save_dir = os.path.expanduser("~")

        # Shared settings variables (created once, used by both Report and Settings pages)
        self.ai_model_var = tk.StringVar(value="deepseek-chat")
        self.api_key_var = tk.StringVar()
        self.column_order_var = tk.StringVar(value="1,2,3,4,5,6,7")
        self.key_issue_highlight_var = tk.BooleanVar(value=True)
        self.comment_timestamp_prefix_var = tk.BooleanVar(value=False)
        self.header_align_var = tk.StringVar(value="left")
        self.cell_align_var = tk.StringVar(value="center")

        self.load_credentials()
        self.setup_ui()

        # Apply saved values AFTER setup_ui creates the vars
        if self.saved_username:
            self.username_var.set(self.saved_username)
            self.password_var.set(self.saved_password)
            self.remember_var.set(True)

        self.api_key_var.set(self.saved_deepseek_api_key)
        self.ai_model_var.set(self.saved_ai_model)
        self.column_order_var.set(self._normalize_column_order(self.saved_column_order))
        self.key_issue_highlight_var.set(self.saved_key_issue_highlight)
        self.comment_timestamp_prefix_var.set(self.saved_comment_timestamp_prefix)

        self.on_key_issue_highlight_toggle()
        self.on_comment_timestamp_toggle()
        self.on_fetch_comment_toggle()

    def load_credentials(self):
        self.saved_username = ""
        self.saved_password = ""
        self.saved_deepseek_api_key = ""
        self.saved_ai_model = "deepseek-chat"
        self.saved_column_order = "1,2,3,4,5,6,7"
        self.saved_key_issue_highlight = True
        self.saved_comment_timestamp_prefix = False
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, "r") as f:
                    data = json.load(f)
                    self.saved_username = data.get("username", "")
                    self.saved_password = data.get("password", "")
                    self.saved_deepseek_api_key = data.get("deepseek_api_key", "")
                    self.saved_ai_model = data.get("ai_model", "deepseek-chat")
                    self.saved_column_order = self._normalize_column_order(data.get("column_order", "1,2,3,4,5,6,7"))
                    self.saved_key_issue_highlight = bool(data.get("key_issue_highlight", True))
                    self.saved_comment_timestamp_prefix = bool(data.get("comment_timestamp_prefix", False))
                    self.last_save_dir = data.get("last_save_dir", os.path.expanduser("~"))
            except:
                pass

    def save_credentials(self, username, password):
        try:
            with open(self.config_file, "w") as f:
                json.dump({
                    "username": username,
                    "password": password,
                    "deepseek_api_key": self.api_key_var.get(),
                    "ai_model": self.ai_model_var.get(),
                    "column_order": self._normalize_column_order(self.column_order_var.get()),
                    "key_issue_highlight": bool(self.key_issue_highlight_var.get()),
                    "comment_timestamp_prefix": bool(self.comment_timestamp_prefix_var.get()),
                    "last_save_dir": self.last_save_dir
                }, f)
        except:
            pass

    def save_ui_preferences(self):
        """Persist non-login UI preferences without touching credentials."""
        try:
            data = {}
            if os.path.exists(self.config_file):
                with open(self.config_file, "r") as f:
                    data = json.load(f)
            data["ai_model"] = self.ai_model_var.get()
            data["column_order"] = self._normalize_column_order(self.column_order_var.get())
            data["key_issue_highlight"] = bool(self.key_issue_highlight_var.get())
            data["comment_timestamp_prefix"] = bool(self.comment_timestamp_prefix_var.get())
            data["last_save_dir"] = self.last_save_dir
            with open(self.config_file, "w") as f:
                json.dump(data, f)
        except:
            pass

    def _normalize_column_order(self, value):
        default = "1,2,3,4,5,6,7"
        text = str(value or "").strip()
        try:
            nums = [int(x.strip()) for x in text.split(",")]
            if len(nums) != 7:
                return default
            if set(nums) != set(range(1, 8)):
                return default
            return ",".join(str(x) for x in nums)
        except:
            return default

    def _field_to_text(self, value):
        """Convert Jira field values (string/dict/list/cascading select) to display text."""
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, dict):
            child = value.get("child")
            child_text = self._field_to_text(child) if child is not None else ""
            val_text = str(value.get("value", "") or "").strip()
            key_text = str(value.get("key", "") or "").strip()
            if child_text and val_text:
                return f"{val_text} - {child_text}"
            return child_text or val_text or key_text
        if isinstance(value, list):
            parts = [self._field_to_text(item) for item in value]
            parts = [p for p in parts if p]
            return " - ".join(parts)
        return str(value).strip()

    def _user_identity_values(self, user):
        """Return comparable Jira user identifiers from a REST user object."""
        if isinstance(user, list):
            values = set()
            for item in user:
                values.update(self._user_identity_values(item))
            return values

        if not isinstance(user, dict):
            return set()

        values = set()
        for key in ("name", "key", "emailAddress", "accountId", "displayName"):
            raw = str(user.get(key, "") or "").strip().lower()
            if raw:
                values.add(raw)
                if "@" in raw:
                    values.add(raw.split("@", 1)[0])
        return values

    def _current_user_identity_values(self):
        values = set()
        for raw in (self.username, self.user_email):
            text = str(raw or "").strip().lower()
            if text:
                values.add(text)
                if "@" in text:
                    values.add(text.split("@", 1)[0])
        return values

    def _is_current_user(self, user):
        return bool(self._current_user_identity_values() & self._user_identity_values(user))

    def _is_jira_issue_key_text(self, text):
        return bool(re.fullmatch(r'[A-Z][A-Z0-9]+-\d+', str(text or "").strip(), flags=re.IGNORECASE))

    def _extract_epic_display_name(self, fields):
        """Extract readable epic title; do not fallback to raw issue key."""
        epic_name = self._field_to_text(fields.get("customfield_10102"))
        if epic_name and not self._is_jira_issue_key_text(epic_name):
            return epic_name

        epic_link = fields.get("customfield_10100")
        if isinstance(epic_link, dict):
            for key in ("summary", "name", "value"):
                text = str(epic_link.get(key, "") or "").strip()
                if text and not self._is_jira_issue_key_text(text):
                    return text

        text = self._field_to_text(epic_link)
        if text and not self._is_jira_issue_key_text(text):
            return text
        return ""

    def _resolve_customer_and_model(self, issue_key, fields):
        """Resolve export values with fallback for R&D issues.

        Normal issue:
          Customer: customfield_11029
          Model:    customfield_12031
        R&D issue (e.g. SW-*):
          Customer: Epic Name(customfield_10102) -> Epic Link(customfield_10100)
          Model:    Platform(customfield_10400/10401)
        """
        issue_key = str(issue_key or "")
        is_rd_issue = issue_key.upper().startswith("SW")

        customer = self._field_to_text(fields.get("customfield_11029"))
        model = self._field_to_text(fields.get("customfield_12031"))

        epic_name = self._extract_epic_display_name(fields)
        platform = self._field_to_text(fields.get("customfield_10400")) or self._field_to_text(fields.get("customfield_10401"))

        if is_rd_issue:
            customer = epic_name or customer
            model = platform or model
        else:
            if not customer:
                customer = epic_name
            if not model:
                model = platform

        return customer, model

    def style_widgets(self):
        """Apply Modern AI Tool style to ttk widgets"""
        style = ttk.Style()
        style.theme_use('clam')

        # Frame
        style.configure("TFrame", background=THEME_SURFACE)

        # Labelframe
        style.configure("TLabelframe", background=THEME_SURFACE, foreground=THEME_TEXT_SECONDARY,
                       bordercolor=THEME_BORDER, relief="flat")
        style.configure("TLabelframe.Label", background=THEME_SURFACE, foreground=THEME_TEXT_SECONDARY,
                       font=("Consolas", 11, "bold"))

        # Primary Button
        style.configure("Modern.TButton", background=THEME_PRIMARY, foreground=THEME_PRIMARY_TEXT,
                       borderwidth=0, relief="flat", font=("Consolas", 12, "bold"), padding=(20, 10))
        style.map("Modern.TButton",
                 background=[("active", THEME_PRIMARY_HOVER), ("pressed", THEME_PRIMARY)],
                 foreground=[("active", THEME_PRIMARY_TEXT)])

        # Secondary Button
        style.configure("Secondary.TButton", background=THEME_SURFACE_RAISED, foreground=THEME_TEXT_SECONDARY,
                       borderwidth=1, bordercolor=THEME_BORDER, relief="flat", font=("Consolas", 11), padding=(16, 8))
        style.map("Secondary.TButton",
                 background=[("active", THEME_SURFACE_HOVER), ("pressed", THEME_SURFACE_RAISED)])

        # Entry
        style.configure("Modern.TEntry", fieldbackground=THEME_SURFACE_RAISED,
                       foreground=THEME_TEXT, bordercolor=THEME_BORDER,
                       borderwidth=1, relief="solid")

        # Combobox
        style.configure("Modern.TCombobox", fieldbackground=THEME_SURFACE_RAISED,
                       foreground=THEME_TEXT, background=THEME_SURFACE_RAISED,
                       bordercolor=THEME_BORDER, borderwidth=1, relief="solid")
        style.map("Modern.TCombobox",
                 fieldbackground=[("readonly", THEME_SURFACE_RAISED)],
                 selectbackground=[("readonly", THEME_PRIMARY)],
                 selectforeground=[("readonly", THEME_TEXT)])

        # Checkbutton
        style.configure("Modern.TCheckbutton", background=THEME_SURFACE,
                       foreground=THEME_TEXT_SECONDARY, font=("Consolas", 11))
        style.map("Modern.TCheckbutton",
                 background=[("active", THEME_SURFACE)],
                 indicatorcolor=[("selected", THEME_PRIMARY), ("!selected", THEME_BORDER)])

        # Scrollbar
        style.configure("Vertical.TScrollbar", background=THEME_BORDER,
                       troughcolor=THEME_SURFACE, bordercolor=THEME_BORDER,
                       arrowcolor=THEME_TEXT_SECONDARY)

        # Progressbar
        style.configure("Modern.Horizontal.TProgressbar",
                       troughcolor=THEME_SURFACE, background=THEME_PRIMARY,
                       bordercolor=THEME_BORDER, lightcolor=THEME_PRIMARY,
                       darkcolor=THEME_PRIMARY_HOVER)

    def _create_card(self, parent, highlight=False):
        """Create a modern card with subtle border and optional highlight"""
        card = tk.Frame(parent, bg=THEME_SURFACE, padx=18, pady=16,
                        highlightbackground=THEME_BORDER, highlightthickness=1,
                        highlightcolor=THEME_BORDER)
        if highlight:
            card.configure(highlightbackground=THEME_PRIMARY, highlightthickness=1,
                          highlightcolor=THEME_PRIMARY)
        return card

    def _add_card_title(self, card, title, subtitle=None):
        """Add title and optional subtitle to a card"""
        title_frame = tk.Frame(card, bg=THEME_SURFACE)
        title_frame.pack(anchor=tk.W, pady=(0, 12))
        tk.Label(title_frame, text=title, font=("Consolas", 13, "bold"),
                fg=THEME_TEXT, bg=THEME_SURFACE).pack(side=tk.LEFT)
        if subtitle:
            tk.Label(title_frame, text=subtitle, font=("Consolas", 9),
                    fg=THEME_TEXT_MUTED, bg=THEME_SURFACE).pack(side=tk.LEFT, padx=(8, 0), pady=(2, 0))

    def _create_section_label(self, parent, text):
        """Create a small section label"""
        return tk.Label(parent, text=text, font=("Consolas", 10),
                       fg=THEME_TEXT_SECONDARY, bg=THEME_SURFACE)

    def _create_input(self, parent, text_var, width=None, show=None):
        """Create a modern input field"""
        kwargs = {"textvariable": text_var, "style": "Modern.TEntry", "font": ("Consolas", 11)}
        if width:
            kwargs["width"] = width
        if show:
            kwargs["show"] = show
        return ttk.Entry(parent, **kwargs)

    def _create_combo(self, parent, text_var, values, width=15):
        """Create a modern combobox"""
        combo = ttk.Combobox(parent, textvariable=text_var, width=width,
                            state="readonly", style="Modern.TCombobox", font=("Consolas", 11))
        combo["values"] = values
        return combo

    def _create_checkbox(self, parent, text, variable, command=None, color=None):
        """Create a modern checkbox"""
        fg = color if color else THEME_TEXT
        cb = tk.Checkbutton(parent, text=text, variable=variable,
                          bg=THEME_SURFACE, fg=fg,
                          selectcolor=THEME_PRIMARY, activebackground=THEME_SURFACE,
                          activeforeground=fg, font=("Consolas", 11),
                          cursor="hand2", command=command)
        return cb

    def _create_primary_button(self, parent, text, command, width=None):
        """Create a primary CTA button with glow effect"""
        btn = tk.Button(parent, text=text, command=command,
                       bg=THEME_PRIMARY, fg=THEME_PRIMARY_TEXT,
                       activebackground=THEME_PRIMARY_HOVER, activeforeground=THEME_PRIMARY_TEXT,
                       relief="flat", borderwidth=0,
                       font=("Consolas", 12, "bold"), cursor="hand2",
                       padx=20, pady=8)
        return btn

    def _create_secondary_button(self, parent, text, command):
        """Create a secondary button"""
        btn = tk.Button(parent, text=text, command=command,
                       bg=THEME_SURFACE_RAISED, fg=THEME_TEXT_SECONDARY,
                       activebackground=THEME_SURFACE_HOVER, activeforeground=THEME_TEXT,
                       relief="flat", borderwidth=1,
                       font=("Consolas", 11), cursor="hand2",
                       padx=16, pady=6)
        return btn

    def setup_ui(self):
        self.style_widgets()

        # Root background
        self.root.configure(bg=THEME_BG)

        # === Layout: Left Sidebar + Main Content ===

        # Main container with sidebar
        container = tk.Frame(self.root, bg=THEME_BG)
        container.pack(fill=tk.BOTH, expand=True)

        # --- Left Sidebar ---
        self.sidebar = tk.Frame(container, bg=THEME_ACTIVITY_BAR, width=56)
        self.sidebar.pack(side=tk.LEFT, fill=tk.Y)
        self.sidebar.pack_propagate(False)

        # Sidebar icons container
        sidebar_icons = tk.Frame(self.sidebar, bg=THEME_ACTIVITY_BAR)
        sidebar_icons.pack(fill=tk.X, pady=(8, 0))

        # Track active page
        self.current_page = tk.StringVar(value="settings")

        # Report icon button
        self.btn_report = tk.Frame(sidebar_icons, bg=THEME_ACTIVITY_BAR, cursor="hand2")
        self.btn_report.pack(pady=(0, 4))
        self.lbl_report = tk.Label(self.btn_report, text="R", font=("Consolas", 15, "bold"),
                bg=THEME_ACTIVITY_BAR, fg="#888888")
        self.lbl_report.pack(padx=12, pady=8)
        self.btn_report.bind("<Button-1>", lambda e: self._show_page("report"))
        self.lbl_report.bind("<Button-1>", lambda e: self._show_page("report"))

        # Login/Settings icon button
        self.btn_settings = tk.Frame(sidebar_icons, bg=THEME_ACTIVITY_BAR, cursor="hand2")
        self.btn_settings.pack(pady=(0, 4))
        self.lbl_settings = tk.Label(self.btn_settings, text="U", font=("Consolas", 15, "bold"),
                bg=THEME_ACTIVITY_BAR, fg="#888888")
        self.lbl_settings.pack(padx=12, pady=8)
        self.btn_settings.bind("<Button-1>", lambda e: self._show_page("settings"))
        self.lbl_settings.bind("<Button-1>", lambda e: self._show_page("settings"))

        # User status at bottom
        self.login_status_sidebar = tk.Label(
            self.sidebar, text="●", font=("Consolas", 10),
            fg=THEME_ERROR, bg=THEME_ACTIVITY_BAR,
            anchor=tk.CENTER
        )
        self.login_status_sidebar.pack(side=tk.BOTTOM, pady=(0, 12))

        # === Page: Report (Main) ===
        self.page_report = tk.Frame(container, bg=THEME_BG)
        self.page_settings = tk.Frame(container, bg=THEME_BG)

        # Settings page will be populated in setup_ui after main_content
        self._setup_settings_page()

        # Show settings page by default (login page)
        self._show_page("settings")

        # === Main Content Area (Report Page) ===
        main_content = self.page_report

        hero = tk.Frame(main_content, bg=THEME_BG)
        hero.pack(fill=tk.X, pady=(0, 16))
        tk.Label(hero, text="Jira Report Studio",
                font=("Consolas", 22, "bold"),
                fg=THEME_TEXT, bg=THEME_BG).pack(anchor=tk.W)
        tk.Label(hero, text="Compose weekly Jira reports with focused filters, comment capture, and optional AI summaries.",
                font=("Consolas", 10),
                fg=THEME_TEXT_SECONDARY, bg=THEME_BG).pack(anchor=tk.W, pady=(4, 0))

        work_area = tk.Frame(main_content, bg=THEME_BG)
        work_area.pack(fill=tk.BOTH, expand=True)
        left_col = tk.Frame(work_area, bg=THEME_BG)
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))
        right_col = tk.Frame(work_area, bg=THEME_BG, width=320)
        right_col.pack(side=tk.LEFT, fill=tk.BOTH, padx=(8, 0))
        right_col.pack_propagate(False)

        # === Card: Date Range ===
        date_card = self._create_card(left_col)
        date_card.pack(fill=tk.X, pady=(0, 12))

        self._add_card_title(date_card, "Date Range", "Select reporting period")

        date_row = tk.Frame(date_card, bg=THEME_SURFACE)
        date_row.pack(fill=tk.X)

        self.start_date_var = tk.StringVar()
        self.end_date_var = tk.StringVar()

        ttk.Entry(date_row, textvariable=self.start_date_var, width=12,
                 style="Modern.TEntry", font=("Consolas", 11)).pack(side=tk.LEFT, fill=tk.X, expand=True)

        tk.Label(date_row, text="—", fg=THEME_TEXT_MUTED, bg=THEME_SURFACE,
                font=("Consolas", 12)).pack(side=tk.LEFT, padx=8)

        ttk.Entry(date_row, textvariable=self.end_date_var, width=12,
                 style="Modern.TEntry", font=("Consolas", 11)).pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Quick date buttons
        quick_row = tk.Frame(date_card, bg=THEME_SURFACE)
        quick_row.pack(fill=tk.X, pady=(10, 0))

        for txt, cmd in [("This Week", lambda: self.set_quick_date("week")),
                        ("Last Week", lambda: self.set_quick_date("last_week")),
                        ("This Month", lambda: self.set_quick_date("month"))]:
            btn = tk.Button(quick_row, text=txt, command=cmd,
                           bg=THEME_SURFACE_RAISED, fg=THEME_TEXT_SECONDARY,
                           activebackground=THEME_SURFACE_HOVER, activeforeground=THEME_TEXT,
                           relief="flat", borderwidth=1,
                           font=("Consolas", 10), cursor="hand2", padx=10, pady=4)
            btn.pack(side=tk.LEFT, padx=(0, 6))

        # Default dates
        today = datetime.date.today()
        week_ago = today - timedelta(days=7)
        self.start_date_var.set(week_ago.strftime("%Y-%m-%d"))
        self.end_date_var.set(today.strftime("%Y-%m-%d"))

        # Auto-update filepath when dates change
        self.start_date_var.trace_add("write", lambda *_: self._update_filepath())
        self.end_date_var.trace_add("write", lambda *_: self._update_filepath())

        # === Card: Filters ===
        filters_card = self._create_card(left_col)
        filters_card.pack(fill=tk.X, pady=(0, 12))

        self._add_card_title(filters_card, "Filters", "Configure report options")

        # Status
        status_row = tk.Frame(filters_card, bg=THEME_SURFACE)
        status_row.pack(fill=tk.X, pady=(0, 10))

        self._create_section_label(status_row, "Status").pack(side=tk.LEFT)
        self.status_filter_var = tk.StringVar(value="ALL")
        self._create_combo(status_row, self.status_filter_var,
                           ["ALL", "WAIT FAE INFO", "WORKED AROUND", "WORKING",
                            "CLOSED", "RESOLVED", "WAIT 3RD PARTY"], width=18).pack(side=tk.LEFT, padx=(8, 0), fill=tk.X, expand=True)

        # Columns
        col_row = tk.Frame(filters_card, bg=THEME_SURFACE)
        col_row.pack(fill=tk.X)

        self._create_section_label(col_row, "Columns").pack(side=tk.LEFT)
        ttk.Entry(col_row, textvariable=self.column_order_var,
                 style="Modern.TEntry", font=("Consolas", 11)).pack(side=tk.LEFT, padx=(8, 0), fill=tk.X, expand=True)

        # === Card: Output ===
        output_card = self._create_card(left_col)
        output_card.pack(fill=tk.X, pady=(0, 12))

        self._add_card_title(output_card, "Output", "Save report to file")

        output_row = tk.Frame(output_card, bg=THEME_SURFACE)
        output_row.pack(fill=tk.X)

        self.filepath_var = tk.StringVar()
        self.filepath_var.set(os.path.join(self.last_save_dir,
                         f"Jira_Weekly_Report_{datetime.date.today().strftime('%Y%m%d')}.xlsx"))
        ttk.Entry(output_row, textvariable=self.filepath_var,
                 style="Modern.TEntry", font=("Consolas", 11)).pack(side=tk.LEFT, fill=tk.X, expand=True)

        tk.Button(output_row, text="Browse", command=self.browse_file,
                 bg=THEME_SURFACE_RAISED, fg=THEME_TEXT,
                 activebackground=THEME_SURFACE_HOVER, relief="flat",
                 borderwidth=1,
                 font=("Consolas", 10), cursor="hand2", padx=10, pady=4).pack(side=tk.LEFT, padx=(8, 0))

        # === Alignment Options ===
        align_card = self._create_card(left_col)
        align_card.pack(fill=tk.X, pady=(0, 16))

        align_row = tk.Frame(align_card, bg=THEME_SURFACE)
        align_row.pack(fill=tk.X)

        self._create_section_label(align_row, "Header").pack(side=tk.LEFT)
        self._create_combo(align_row, self.header_align_var, ["left", "center", "right"], width=8).pack(side=tk.LEFT, padx=(4, 16))

        self._create_section_label(align_row, "Cell").pack(side=tk.LEFT)
        self._create_combo(align_row, self.cell_align_var, ["left", "center", "right"], width=8).pack(side=tk.LEFT, padx=(4, 16), fill=tk.X, expand=True)

        self._create_checkbox(align_row, "Key Issue Red",
                             self.key_issue_highlight_var, color=THEME_ERROR).pack(side=tk.LEFT)

        # === Card: Progress Content ===
        progress_card = self._create_card(right_col)
        progress_card.pack(fill=tk.X, pady=(0, 12))

        self._add_card_title(progress_card, "Progress Content", "Non-AI progress source")

        self.fetch_comment_var = tk.BooleanVar(value=False)
        self._create_checkbox(progress_card, "Fetch latest comment",
                             self.fetch_comment_var, command=self.on_fetch_comment_toggle).pack(anchor=tk.W, pady=(0, 8))

        self._create_checkbox(progress_card, "Prefix timestamp",
                             self.comment_timestamp_prefix_var, command=self.on_comment_timestamp_toggle).pack(anchor=tk.W)

        # === Card: AI Summary (Highlighted) ===
        ai_card = self._create_card(right_col, highlight=True)
        ai_card.pack(fill=tk.X, pady=(0, 12))

        # AI card header with glow effect
        ai_header = tk.Frame(ai_card, bg=THEME_SURFACE)
        ai_header.pack(fill=tk.X, pady=(0, 12))

        tk.Label(ai_header, text="AI Summary", font=("Consolas", 13, "bold"),
                fg=THEME_PRIMARY, bg=THEME_SURFACE).pack(side=tk.LEFT)
        tk.Label(ai_header, text="Powered by DeepSeek", font=("Consolas", 9),
                fg=THEME_TEXT_MUTED, bg=THEME_SURFACE).pack(side=tk.LEFT, padx=(8, 0), pady=(2, 0))

        # AI options
        self.use_ai_summary_var = tk.BooleanVar(value=False)
        self._create_checkbox(ai_card, "Enable AI-powered summary",
                             self.use_ai_summary_var, command=self.on_ai_summary_toggle,
                             color=THEME_PRIMARY).pack(anchor=tk.W, pady=(0, 8))

        # AI Config (nested)
        self.ai_config_outer = tk.Frame(ai_card, bg=THEME_SURFACE_RAISED, padx=10, pady=8)
        self.ai_config_outer.pack(fill=tk.X, pady=(0, 8))

        model_row = tk.Frame(self.ai_config_outer, bg=THEME_SURFACE_RAISED)
        model_row.pack(fill=tk.X, pady=(0, 6))

        self._create_section_label(model_row, "Model").pack(side=tk.LEFT)
        self._create_combo(model_row, self.ai_model_var,
                           ["deepseek-chat", "deepseek-coder", "deepseek-v4-flash", "deepseek-v4-pro"],
                           width=18).pack(side=tk.LEFT, padx=(8, 0), fill=tk.X, expand=True)

        batch_row = tk.Frame(self.ai_config_outer, bg=THEME_SURFACE_RAISED)
        batch_row.pack(fill=tk.X)

        self.batch_mode_var = tk.BooleanVar(value=False)
        self._create_checkbox(batch_row, "Batch Mode", self.batch_mode_var,
                             command=self.on_batch_mode_toggle).pack(side=tk.LEFT)

        tk.Label(batch_row, text="Size", font=("Consolas", 10),
                fg=THEME_TEXT_SECONDARY, bg=THEME_SURFACE_RAISED).pack(side=tk.LEFT, padx=(12, 4))
        self.batch_size_var = tk.IntVar(value=10)
        tk.Entry(batch_row, textvariable=self.batch_size_var, width=4,
                bg=THEME_SURFACE, fg=THEME_TEXT, insertbackground=THEME_TEXT,
                relief="solid", bd=1, font=("Consolas", 10), justify=tk.CENTER).pack(side=tk.LEFT)

        self.ai_config_outer.pack_forget()

        # === Action Buttons (Sidebar) ===
        sidebar_actions = tk.Frame(self.sidebar, bg=THEME_ACTIVITY_BAR)
        sidebar_actions.pack(side=tk.BOTTOM, fill=tk.X, pady=(8, 8))

        self.generate_btn = self._create_primary_button(
            sidebar_actions, "▶", self.generate_report
        )
        self.generate_btn.pack(fill=tk.X, padx=6, pady=(0, 4))
        self.generate_btn.config(state=tk.DISABLED, width=3)

        self.cancel_btn = self._create_secondary_button(
            sidebar_actions, "✕", self.cancel_generation
        )
        self.cancel_btn.pack(fill=tk.X, padx=6)
        self.cancel_btn.config(state=tk.DISABLED, width=3)

        # === Sidebar Progress Fill (Full Background Animation) ===
        self.progress_fill_frame = tk.Frame(self.sidebar, bg=THEME_ACTIVITY_BAR)
        self.progress_fill_frame.pack(side=tk.LEFT, fill=tk.Y, ipadx=0)
        self.progress_fill = tk.Frame(self.progress_fill_frame, bg=THEME_PRIMARY)
        self.progress_fill.place(relx=0, rely=1, relw=1, relh=0, anchor="sw")
        self.sidebar_progress = 0

        # === Processing Indicator ===
        self.processing_frame = tk.Frame(main_content, bg=THEME_BG)
        self.processing_frame.pack_forget()

        self.spinner_label = tk.Label(self.processing_frame, text="◐",
                                     font=("Consolas", 16), fg=THEME_PRIMARY, bg=THEME_BG)
        self.spinner_label.pack(side=tk.LEFT, padx=(0, 8))

        self.processing_status = tk.Label(self.processing_frame, text="",
                                         font=("Consolas", 11), fg=THEME_TEXT, bg=THEME_BG, anchor=tk.W)
        self.processing_status.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.processing_detail = tk.Label(self.processing_frame, text="",
                                        font=("Consolas", 9), fg=THEME_TEXT_MUTED, bg=THEME_BG, anchor=tk.W)
        self.processing_detail.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(main_content, variable=self.progress_var,
                                          maximum=100, mode="determinate",
                                          style="Modern.Horizontal.TProgressbar")
        self.progress_bar.pack(fill=tk.X, pady=(0, 8))
        self.progress_bar.pack_forget()

        self.spinner_frames = ["◐", "◓", "◑", "◒"]
        self.spinner_index = 0
        self.spinner_running = False

        # === Status Bar ===
        self.status_bar = tk.Label(self.root, text="Ready — Sign in to continue",
                                   bd=0, relief=tk.FLAT, anchor=tk.W, padx=8,
                                   font=("Consolas", 10), fg=THEME_TEXT_SECONDARY, bg=THEME_STATUS_BAR)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def set_quick_date(self, period):
        today = datetime.date.today()
        if period == "week":
            days_since_monday = today.weekday()
            start = today - timedelta(days=days_since_monday)
            end = start + timedelta(days=6)
        elif period == "last_week":
            days_since_monday = today.weekday()
            this_monday = today - timedelta(days=days_since_monday)
            last_monday = this_monday - timedelta(days=7)
            last_sunday = last_monday + timedelta(days=6)
            start = last_monday
            end = last_sunday
        else:
            start = today.replace(day=1)
            _, last_day = calendar.monthrange(today.year, today.month)
            end = today.replace(day=last_day)
        self.start_date_var.set(start.strftime("%Y-%m-%d"))
        self.end_date_var.set(end.strftime("%Y-%m-%d"))
        self._update_filepath()

    def _update_filepath(self):
        """Update the output filepath based on current dates and username."""
        start_str = self.start_date_var.get()
        end_str = self.end_date_var.get()
        # Extract username before @ for filename
        username_short = self.username.split("@")[0] if self.username else "jira_report"
        filename = f"{username_short}_{start_str}_{end_str}_jira_report.xlsx"
        self.filepath_var.set(os.path.join(self.last_save_dir, filename))

    def browse_file(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=self.last_save_dir,
            initialfile=os.path.basename(self.filepath_var.get())
        )
        if file_path:
            self.filepath_var.set(file_path)
            self.last_save_dir = os.path.dirname(file_path)
            self.save_ui_preferences()

    def toggle_password_visibility(self):
        if self.show_password_var.get():
            self.password_entry.config(show="")
        else:
            self.password_entry.config(show="*")

    def _show_page(self, page_name):
        """Switch between pages (report/settings)"""
        self.current_page.set(page_name)

        # Hide all pages
        self.page_report.pack_forget()
        self.page_settings.pack_forget()

        # Show selected page
        if page_name == "report":
            self.page_report.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=16, pady=12)
        else:
            self.page_settings.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=16, pady=12)

        self._highlight_sidebar(page_name)

    def _highlight_sidebar(self, active):
        """Highlight the active sidebar button"""
        bg_active = "#1e1e1e"
        fg_active = "#ffffff"
        bg_inactive = THEME_ACTIVITY_BAR
        fg_inactive = "#888888"

        # Reset all
        self.btn_report.configure(bg=bg_inactive)
        self.btn_settings.configure(bg=bg_inactive)
        for child in self.btn_report.winfo_children():
            child.configure(bg=bg_inactive, fg=fg_inactive)
        for child in self.btn_settings.winfo_children():
            child.configure(bg=bg_inactive, fg=fg_inactive)

        # Highlight active
        if active == "report":
            self.btn_report.configure(bg=bg_active)
            for child in self.btn_report.winfo_children():
                child.configure(bg=bg_active, fg=THEME_PRIMARY)
        else:
            self.btn_settings.configure(bg=bg_active)
            for child in self.btn_settings.winfo_children():
                child.configure(bg=bg_active, fg=THEME_PRIMARY)

    def _setup_settings_page(self):
        """Setup the Settings page with configuration options"""
        # Title
        tk.Label(self.page_settings, text="Settings",
                font=("Consolas", 18, "bold"),
                fg=THEME_TEXT, bg=THEME_BG).pack(anchor=tk.W, pady=(0, 16))

        # === Jira Connection Card ===
        conn_card = self._create_card(self.page_settings)
        conn_card.pack(fill=tk.X, pady=(0, 12))

        self._add_card_title(conn_card, "Jira Connection")

        # Username
        row = tk.Frame(conn_card, bg=THEME_SURFACE)
        row.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row, text="Username:", font=("Consolas", 10),
                fg=THEME_TEXT_SECONDARY, bg=THEME_SURFACE).pack(side=tk.LEFT)
        self.username_var = tk.StringVar(value="")
        ttk.Entry(row, textvariable=self.username_var,
                 style="Modern.TEntry", font=("Consolas", 10)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 0))

        # Password
        row = tk.Frame(conn_card, bg=THEME_SURFACE)
        row.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row, text="Password:", font=("Consolas", 10),
                fg=THEME_TEXT_SECONDARY, bg=THEME_SURFACE).pack(side=tk.LEFT)
        self.password_var = tk.StringVar(value="")
        self.password_entry = ttk.Entry(row, textvariable=self.password_var, show="*",
                                        style="Modern.TEntry", font=("Consolas", 10))
        self.password_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 0))

        # Remember checkbox
        self.remember_var = tk.BooleanVar(value=False)
        tk.Checkbutton(conn_card, text="Remember credentials",
                      variable=self.remember_var,
                      bg=THEME_SURFACE, fg=THEME_TEXT_SECONDARY,
                      selectcolor=THEME_PRIMARY, activebackground=THEME_SURFACE,
                      font=("Consolas", 9), cursor="hand2").pack(anchor=tk.W, pady=(0, 8))

        # Login buttons
        btn_row = tk.Frame(conn_card, bg=THEME_SURFACE)
        btn_row.pack(fill=tk.X)
        self.login_btn = ttk.Button(btn_row, text="Login", command=self.login,
                                   width=10, style="Modern.TButton")
        self.login_btn.pack(side=tk.LEFT)
        self.logout_btn = ttk.Button(btn_row, text="Logout", command=self.logout,
                                    state=tk.DISABLED, width=10, style="Secondary.TButton")
        self.logout_btn.pack(side=tk.LEFT, padx=(8, 0))

        self.login_status_label = tk.Label(conn_card, text="Not connected",
                                          font=("Consolas", 9),
                                          fg=THEME_ERROR, bg=THEME_SURFACE)
        self.login_status_label.pack(anchor=tk.W, pady=(8, 0))

        # === AI Settings Card ===
        ai_card = self._create_card(self.page_settings)
        ai_card.pack(fill=tk.X, pady=(0, 12))

        self._add_card_title(ai_card, "AI Settings")

        row = tk.Frame(ai_card, bg=THEME_SURFACE)
        row.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row, text="API Key:", font=("Consolas", 10),
                fg=THEME_TEXT_SECONDARY, bg=THEME_SURFACE).pack(side=tk.LEFT)
        ttk.Entry(row, textvariable=self.api_key_var, show="*",
                 style="Modern.TEntry", font=("Consolas", 10)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 0))

        row = tk.Frame(ai_card, bg=THEME_SURFACE)
        row.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row, text="Model:", font=("Consolas", 10),
                fg=THEME_TEXT_SECONDARY, bg=THEME_SURFACE).pack(side=tk.LEFT)
        ttk.Combobox(row, textvariable=self.ai_model_var, width=18, state="readonly",
                    style="Modern.TCombobox", font=("Consolas", 10),
                    values=["deepseek-chat", "deepseek-coder", "deepseek-v4-flash", "deepseek-v4-pro"]).pack(side=tk.LEFT, padx=(8, 0), fill=tk.X, expand=True)

        # === Default Report Settings Card ===
        report_card = self._create_card(self.page_settings)
        report_card.pack(fill=tk.X, pady=(0, 12))

        self._add_card_title(report_card, "Default Report Settings")

        row = tk.Frame(report_card, bg=THEME_SURFACE)
        row.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row, text="Columns:", font=("Consolas", 10),
                fg=THEME_TEXT_SECONDARY, bg=THEME_SURFACE).pack(side=tk.LEFT)
        ttk.Entry(row, textvariable=self.column_order_var,
                 style="Modern.TEntry", font=("Consolas", 10)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 0))

        row = tk.Frame(report_card, bg=THEME_SURFACE)
        row.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row, text="Header Align:", font=("Consolas", 10),
                fg=THEME_TEXT_SECONDARY, bg=THEME_SURFACE).pack(side=tk.LEFT)
        ttk.Combobox(row, textvariable=self.header_align_var, width=10, state="readonly",
                    style="Modern.TCombobox", font=("Consolas", 10),
                    values=["left", "center", "right"]).pack(side=tk.LEFT, padx=(8, 0))

        row = tk.Frame(report_card, bg=THEME_SURFACE)
        row.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row, text="Cell Align:", font=("Consolas", 10),
                fg=THEME_TEXT_SECONDARY, bg=THEME_SURFACE).pack(side=tk.LEFT)
        ttk.Combobox(row, textvariable=self.cell_align_var, width=10, state="readonly",
                    style="Modern.TCombobox", font=("Consolas", 10),
                    values=["left", "center", "right"]).pack(side=tk.LEFT, padx=(8, 0))

        tk.Checkbutton(report_card, text="Highlight key issues in red",
                      variable=self.key_issue_highlight_var,
                      bg=THEME_SURFACE, fg=THEME_ERROR,
                      selectcolor=THEME_PRIMARY, activebackground=THEME_SURFACE,
                      font=("Consolas", 9), cursor="hand2").pack(anchor=tk.W, pady=(0, 8))

        # Save button
        tk.Button(report_card, text="Save Settings", command=self._save_settings,
                 bg=THEME_PRIMARY, fg=THEME_PRIMARY_TEXT,
                 activebackground=THEME_PRIMARY_HOVER, relief="flat",
                 font=("Consolas", 10, "bold"), cursor="hand2", padx=16, pady=6).pack(anchor=tk.E)

    def _save_settings(self):
        """Save all settings to config file"""
        try:
            data = {
                "username": self.username_var.get(),
                "password": self.password_var.get() if self.remember_var.get() else "",
                "deepseek_api_key": self.api_key_var.get(),
                "ai_model": self.ai_model_var.get(),
                "column_order": self._normalize_column_order(self.column_order_var.get()),
                "key_issue_highlight": bool(self.key_issue_highlight_var.get()),
                "comment_timestamp_prefix": bool(self.comment_timestamp_prefix_var.get()),
                "last_save_dir": self.last_save_dir
            }
            with open(self.config_file, "w") as f:
                json.dump(data, f)
            messagebox.showinfo("Success", "Settings saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save settings:\n{e}")

    def on_ai_summary_toggle(self):
        if self.use_ai_summary_var.get():
            self.ai_config_outer.pack(fill=tk.X, pady=(0, 8))
            self.fetch_comment_var.set(False)
        else:
            self.ai_config_outer.pack_forget()

    def on_fetch_comment_toggle(self):
        if self.fetch_comment_var.get():
            self.use_ai_summary_var.set(False)
            self.ai_config_outer.pack_forget()

    def on_batch_mode_toggle(self):
        self.save_ui_preferences()

    def on_key_issue_highlight_toggle(self):
        self.save_ui_preferences()

    def on_comment_timestamp_toggle(self):
        self.save_ui_preferences()

    def login(self):
        url = self.base_url
        username = self.username_var.get().strip()
        password = self.password_var.get()

        if not username or not password:
            messagebox.showerror("Error", "Please enter username and password")
            return

        self.login_btn.config(state=tk.DISABLED)
        self.update_status("Logging in...")

        import threading
        thread = threading.Thread(target=self._login_thread, args=(url, username, password))
        thread.daemon = True
        thread.start()

    def _login_thread(self, url, username, password):
        try:
            result = self._do_api_login(url, username, password)
            self.root.after(0, lambda: self._handle_login_result(result))
        except Exception as e:
            self.root.after(0, lambda: self._handle_login_error(str(e)))

    def _do_api_login(self, url, username, password):
        try:
            auth = requests.auth.HTTPBasicAuth(username, password)
            response = self.session.get(f"{url}/rest/api/2/myself", auth=auth, timeout=30)

            if response.status_code == 200:
                user_data = response.json()
                self.user_email = user_data.get("email", "")
                self.session.auth = auth
                return {"success": True, "username": username}
            elif response.status_code == 401:
                return self._do_cookie_login(url, username, password)
            else:
                return {"success": False, "error": f"API returned status {response.status_code}"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def _do_cookie_login(self, url, username, password):
        try:
            login_page = self.session.get(f"{url}/login.jsp", timeout=30)
            atl_token_match = re.search(
                r'name="atl_token"\s*type="hidden"\s*value="([^"]+)"',
                login_page.text
            )
            atl_token = atl_token_match.group(1) if atl_token_match else ""

            form_data = {
                "os_username": username,
                "os_password": password,
                "os_destination": "/",
                "atl_token": atl_token,
                "user_role": "",
                "os_cookie": "true"
            }
            login_response = self.session.post(
                f"{url}/dologin.jsp",
                data=form_data,
                timeout=30,
                allow_redirects=True
            )

            if "invalid" in login_response.text.lower() or "incorrect" in login_response.text.lower():
                return {"success": False, "error": "Invalid username or password"}

            api_check = self.session.get(f"{url}/rest/api/2/myself", timeout=30)
            if api_check.status_code == 200:
                user_data = api_check.json()
                self.user_email = user_data.get("email", "")
                return {"success": True, "username": username}
            else:
                return {"success": False, "error": f"Verification failed (status: {api_check.status_code})"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def _handle_login_result(self, result):
        self.login_btn.config(state=tk.NORMAL)
        if result["success"]:
            self.username = result["username"]
            self.on_login_success(self.username)
        else:
            messagebox.showerror("Login Failed", result["error"])
            self.update_status("Login failed")

    def _handle_login_error(self, error):
        self.login_btn.config(state=tk.NORMAL)
        messagebox.showerror("Connection Error", f"Cannot connect to Jira server:\n{error}")
        self.update_status("Connection error")

    def on_login_success(self, username):
        self.logged_in = True
        self.username = username

        if self.remember_var.get():
            self.save_credentials(username, self.password_var.get())

        self.login_status_sidebar.config(text=f"●", fg=THEME_SUCCESS)
        self.login_status_label.config(text=f"Connected: {username}", fg=THEME_SUCCESS)
        self.login_btn.config(state=tk.DISABLED)
        self.logout_btn.config(state=tk.NORMAL)
        self.generate_btn.config(state=tk.NORMAL)
        self._update_filepath()
        self.update_status(f"► Logged in as {username}")
        self._show_page("report")

    def logout(self):
        if self.logged_in:
            try:
                self.session.delete(f"{self.base_url}/rest/auth/1/session")
            except:
                pass
            self.logged_in = False
            self.username = None
            self.user_email = None
            self.login_status_sidebar.config(text="●", fg=THEME_ERROR)
            self.login_status_label.config(text="Not connected", fg=THEME_ERROR)
            self.login_btn.config(state=tk.NORMAL)
            self.logout_btn.config(state=tk.DISABLED)
            self.generate_btn.config(state=tk.DISABLED)
            self.cancel_btn.config(state=tk.DISABLED)
            self.update_status("► Logged out")

    def update_status(self, message):
        self.status_bar.config(text=f"► {message}")
        self.root.update_idletasks()

    def _spin_step(self):
        """Update spinner animation frame"""
        if self.spinner_running:
            self.spinner_label.config(text=self.spinner_frames[self.spinner_index])
            self.spinner_index = (self.spinner_index + 1) % len(self.spinner_frames)
            self.root.after(100, self._spin_step)

    def start_processing(self, status_text="Processing..."):
        """Show processing animation"""
        self.spinner_running = True
        self.spinner_index = 0
        self.processing_status.config(text=status_text)
        self.processing_detail.config(text="")
        self.progress_var.set(0)
        self.processing_frame.pack(fill=tk.X, pady=(8, 0))
        self.progress_bar.pack(fill=tk.X, pady=(0, 8))
        self.sidebar_progress = 0
        self._update_sidebar_progress(0)
        self._spin_step()
        self.root.update_idletasks()

    def _update_sidebar_progress(self, progress):
        """Update sidebar background fill progress (0-100)"""
        self.sidebar_progress = progress
        relh = progress / 100.0
        self.progress_fill.place_configure(relh=relh)
        self.root.update_idletasks()

    def update_processing(self, status_text, detail_text="", progress=None):
        """Update processing status"""
        self.processing_status.config(text=status_text)
        self.processing_detail.config(text=detail_text)
        if progress is not None:
            self.progress_var.set(max(0, min(100, progress)))
            self._update_sidebar_progress(progress)
        self.root.update_idletasks()

    def stop_processing(self):
        """Hide processing animation"""
        self.spinner_running = False
        self.processing_frame.pack_forget()
        self.progress_bar.pack_forget()
        self.progress_var.set(0)
        self._update_sidebar_progress(0)
        self.root.update_idletasks()

    def cancel_generation(self):
        """Request cancellation of the current report generation."""
        if not self.generation_running:
            return
        self.cancel_event.set()
        self.cancel_btn.config(state=tk.DISABLED, text="...")
        self.update_processing("Cancelling...", "Waiting for current request to finish...", self.progress_var.get())

    def check_cancelled(self):
        if self.cancel_event.is_set():
            raise OperationCancelled()

    def finish_generation_ui(self, enable_generate=True):
        self.generation_running = False
        self.cancel_event.clear()
        self.stop_processing()
        if self.logged_in and enable_generate:
            self.generate_btn.config(state=tk.NORMAL)
        self.cancel_btn.config(state=tk.DISABLED, text="✕")

    def generate_report(self):
        if not self.logged_in:
            messagebox.showerror("Error", "Please login first")
            return

        try:
            start_date = datetime.datetime.strptime(self.start_date_var.get(), "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(self.end_date_var.get(), "%Y-%m-%d").date()
        except ValueError:
            messagebox.showerror("Error", "Invalid date format. Use YYYY-MM-DD")
            return

        if end_date < start_date:
            messagebox.showerror("Error", "End date must be after start date")
            return

        selected_status = self.status_filter_var.get().strip()
        if selected_status == "ALL":
            status_clause = ""
        else:
            status_clause = f'status = "{selected_status}" '

        engineer_field = "Software Development Engineer 软件开发工程师"

        filepath = self.filepath_var.get().strip()
        if not filepath:
            messagebox.showerror("Error", "Please select a save path")
            return
        if not filepath.endswith(".xlsx"):
            filepath += ".xlsx"

        save_dir = os.path.dirname(filepath)
        if save_dir and not os.path.exists(save_dir):
            os.makedirs(save_dir)

        self.column_order_var.set(self._normalize_column_order(self.column_order_var.get()))
        self.save_ui_preferences()

        self.cancel_event.clear()
        self.generation_running = True
        self.generate_btn.config(state=tk.DISABLED)
        self.cancel_btn.config(state=tk.NORMAL, text="✕")
        self.start_processing("Starting...")

        thread = threading.Thread(target=self._generate_report_work,
                                  args=(start_date, end_date, selected_status, engineer_field, filepath))
        thread.daemon = True
        thread.start()

    def _generate_report_work(self, start_date, end_date, selected_status, engineer_field, filepath):
        try:
            status_clause = f'status = "{selected_status}" ' if selected_status != "ALL" else ""

            jql_normal = f'"{engineer_field}" IN (currentUser()) AND updated >= {start_date} AND updated <= "{end_date} 23:59"'
            if status_clause:
                jql_normal += f' AND {status_clause}'

            jql_wait3rd = f'"{engineer_field}" IN (currentUser()) AND status = "WAIT 3RD PARTY" AND updated >= {start_date} AND updated <= "{end_date} 23:59"'
            if status_clause:
                jql_wait3rd += f' AND {status_clause}'

            jql_assist_normal = f'comment ~ currentUser() AND "{engineer_field}" != currentUser() AND updated >= {start_date} AND updated <= "{end_date} 23:59"'
            if status_clause:
                jql_assist_normal += f' AND {status_clause}'

            jql_assist_wait3rd = f'comment ~ currentUser() AND "{engineer_field}" != currentUser() AND status = "WAIT 3RD PARTY" AND updated >= {start_date} AND updated <= "{end_date} 23:59"'
            if status_clause:
                jql_assist_wait3rd += f' AND {status_clause}'

            self.check_cancelled()
            self.root.after(0, lambda: self.update_processing("Searching issues...", f"Searching assigned issues...", 5))
            issues_assigned_normal = self.fetch_issues(jql_normal)
            self.check_cancelled()
            self.root.after(0, lambda: self.update_processing("Searching issues...", f"Searching assigned WAIT_3RD issues...", 10))
            issues_assigned_wait3rd = self.fetch_issues(jql_wait3rd)
            issues_assigned = issues_assigned_normal + issues_assigned_wait3rd
            self.check_cancelled()
            self.root.after(0, lambda: self.update_processing(f"Found {len(issues_assigned)} assigned issues", f"{len(issues_assigned_normal)} normal + {len(issues_assigned_wait3rd)} WAIT_3RD", 20))

            issues_assist_normal = self.fetch_issues(jql_assist_normal)
            self.check_cancelled()
            self.root.after(0, lambda: self.update_processing("Searching assist issues...", f"Searching assist WAIT_3RD issues...", 25))
            issues_assist_wait3rd = self.fetch_issues(jql_assist_wait3rd)
            issues_assist = issues_assist_normal + issues_assist_wait3rd
            self.check_cancelled()
            self.root.after(0, lambda: self.update_processing(f"Found {len(issues_assist)} assist issues", f"{len(issues_assist_normal)} normal + {len(issues_assist_wait3rd)} WAIT_3RD", 35))

            no_comment_required_statuses = {"WAIT 3RD PARTY", "WORKING"}
            wait_blocked_statuses = {"WAIT FAE INFO", "WORKED AROUND", "WAIT OFFICIAL RELEASE"}
            closed_statuses = {"CLOSED", "RESOLVED"}

            def is_in_date_range(issue):
                """Check if issue's created time or latest comment time is within date range"""
                created_str = issue.get("fields", {}).get("created", "")
                dt = self._parse_jira_datetime(created_str)
                if dt and start_date <= dt.date() <= end_date:
                    return True
                return self.user_commented_in_date_range(issue['key'], start_date, end_date)

            def field_date_in_range(issue, field_name):
                date_str = issue.get("fields", {}).get(field_name, "")
                dt = self._parse_jira_datetime(date_str)
                return bool(dt and start_date <= dt.date() <= end_date)

            def should_include_issue(issue, start_date, end_date):
                """判断一个 issue 是否应该包含在报告中

                逻辑：
                1. WAIT 3RD PARTY/WORKING：created时间或评论时间在时间区间内才保留
                2. WAIT FAE INFO/WORKED AROUND/WAIT OFFICIAL RELEASE：
                   - 情景1：SDE是当前用户 且 assignee非当前用户 → 直接保留
                   - 情景3：SDE非当前用户 → 当前用户在时间区间内有评论才保留
                3. CLOSED/RESOLVED：SDE是当前用户且本期关闭/解决 → 保留
                4. 其他状态：当前用户在时间区间内有评论就保留
                """
                status = issue.get("fields", {}).get("status", {}).get("name", "")
                status_key = status.upper()
                fields = issue.get("fields", {})
                assignee_field = fields.get("assignee")
                sde_field = fields.get("customfield_12001")

                if status_key in no_comment_required_statuses:
                    return is_in_date_range(issue)

                if status_key in wait_blocked_statuses:
                    # 情景1：SDE是当前用户 且 assignee非当前用户 → 直接保留
                    # 情景3：SDE非当前用户 → 当前用户在时间区间内有评论才保留
                    # JQL 的 comment/currentUser 只能作为粗筛，最终以 REST 字段和评论作者为准。
                    if self._is_current_user(sde_field):
                        if not self._is_current_user(assignee_field):
                            # 情景1：SDE是当前用户 且 assignee非当前用户 → 直接保留
                            return True
                        # assignee是当前用户，需要当前用户在时间区间内有评论
                        return self.user_commented_in_date_range(issue['key'], start_date, end_date)

                    # SDE不是当前用户，是情景3；需要当前用户在时间区间内有评论才保留。
                    return self.user_commented_in_date_range(issue['key'], start_date, end_date)

                if status_key in closed_statuses and self._is_current_user(sde_field):
                    if field_date_in_range(issue, "resolutiondate") or field_date_in_range(issue, "updated"):
                        return True

                # 其他状态：当前用户在时间区间内有评论就保留
                return self.user_commented_in_date_range(issue['key'], start_date, end_date)

            self.root.after(0, lambda: self.update_processing("Filtering issues...", f"Checking {len(issues_assigned)} assigned issues"))
            issues_assigned_filtered = []
            assigned_total = max(len(issues_assigned), 1)
            for idx, issue in enumerate(issues_assigned, 1):
                self.check_cancelled()
                progress = 35 + (idx / assigned_total) * 15
                status = issue.get("fields", {}).get("status", {}).get("name", "")
                if should_include_issue(issue, start_date, end_date):
                    issues_assigned_filtered.append(issue)
                else:
                    self.root.after(0, lambda k=issue['key'], p=progress: self.update_processing("Filtering issues...", f"Skipping {k} - no activity", p))
            issues_assigned = issues_assigned_filtered

            self.check_cancelled()
            self.root.after(0, lambda: self.update_processing("Filtering assist issues...", f"Checking {len(issues_assist)} assist issues", 50))
            issues_assist_filtered = []
            assist_total = max(len(issues_assist), 1)
            for idx, issue in enumerate(issues_assist, 1):
                self.check_cancelled()
                progress = 50 + (idx / assist_total) * 15
                if should_include_issue(issue, start_date, end_date):
                    issues_assist_filtered.append(issue)
                else:
                    self.root.after(0, lambda k=issue['key'], p=progress: self.update_processing("Filtering issues...", f"Skipping {k} - no activity", p))
            issues_assist = issues_assist_filtered

            all_issues = {issue['key']: issue for issue in issues_assigned + issues_assist}
            issues = list(all_issues.values())

            def get_created_timestamp(issue):
                created_str = issue.get("fields", {}).get("created", "")
                dt = self._parse_jira_datetime(created_str)
                return dt.timestamp() if dt else 0

            issues.sort(key=lambda x: -get_created_timestamp(x))

            self.check_cancelled()
            self.root.after(0, lambda: self.update_processing(f"Found {len(issues)} total issues", "Generating Excel...", 70))
            self.root.after(0, lambda: self.update_processing("Generating Excel file...", "Please wait...", 72))

            self.create_excel(issues, filepath, selected_status, start_date, end_date)
            self.check_cancelled()

            self.root.after(0, lambda: self.update_processing("Done", "Report saved", 100))
            self.root.after(0, lambda: self.finish_generation_ui())
            self.root.after(0, lambda: self.update_status(f"Report saved: {filepath}"))
            self.root.after(0, lambda: messagebox.showinfo("Success", f"Report generated successfully!\n\n{len(issues)} issues exported to:\n{filepath}"))

        except OperationCancelled:
            self.root.after(0, lambda: self.finish_generation_ui())
            self.root.after(0, lambda: self.update_status("Cancelled"))
            self.root.after(0, lambda: messagebox.showinfo("Cancelled", "Report generation was cancelled."))
        except Exception as e:
            self.root.after(0, lambda: self.finish_generation_ui())
            self.root.after(0, lambda: messagebox.showerror("Error", f"Failed:\n{str(e)}"))
            self.root.after(0, lambda: self.update_status("Failed"))

    def fetch_issues(self, jql, start_at=0, max_results=100):
        all_issues = []
        url = f"{self.base_url}/rest/api/2/search"

        params = {
            "jql": jql,
            "startAt": start_at,
            "maxResults": max_results,
            "fields": "summary,status,priority,created,updated,resolutiondate,creator,key,assignee,customfield_12001,customfield_11029,customfield_12031,customfield_10100,customfield_10102,customfield_10400,customfield_10401"
        }

        while True:
            self.check_cancelled()
            try:
                response = self.session.get(url, params=params, timeout=30)
                self.check_cancelled()

                if response.status_code >= 400:
                    error_detail = response.text[:500] if response.text else "No details"
                    raise Exception(f"Error {response.status_code}: {error_detail}")

                response.raise_for_status()
                data = response.json()

                issues = data.get("issues", [])
                all_issues.extend(issues)

                total = data.get("total", 0)
                if start_at + len(issues) >= total:
                    break

                start_at += max_results
                self.root.after(0, lambda c=len(all_issues), t=total: self.update_status(f"Fetching {c}/{t}..."))

            except requests.exceptions.RequestException as e:
                raise Exception(f"Fetch error: {str(e)}")

        return all_issues

    def _parse_jira_datetime(self, created_str):
        """Parse Jira timestamp like 2026-05-08T17:32:01.123+0800 into datetime."""
        if not created_str:
            return None
        raw = str(created_str).strip()
        if not raw:
            return None

        # Normalize timezone suffix +0800 -> +08:00 for fromisoformat
        if re.search(r'[+-]\d{4}$', raw):
            raw = raw[:-5] + raw[-5:-2] + ":" + raw[-2:]

        try:
            return datetime.datetime.fromisoformat(raw)
        except ValueError:
            pass

        for fmt in ("%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.datetime.strptime(raw, fmt)
            except ValueError:
                continue
        return None

    def _clean_comment_for_display(self, body):
        """Light cleanup for 'Fetch latest comment' to preserve user wording."""
        if not body:
            return ""

        text = body
        text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
        text = re.sub(r'</(?:p|li|div|tr)>', '\n', text, flags=re.IGNORECASE)
        text = re.sub(r'<[^>]+>', '', text)
        text = (text.replace('&nbsp;', ' ').replace('&amp;', '&')
                    .replace('&lt;', '<').replace('&gt;', '>')
                    .replace('&quot;', '"').replace('&#39;', "'"))
        text = re.sub(r'[\r\n]+', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()
        return text

    def user_commented_in_date_range(self, issue_key, start_date, end_date):
        try:
            self.check_cancelled()
            url = f"{self.base_url}/rest/api/2/issue/{issue_key}/comment"
            response = self.session.get(url, timeout=30)
            self.check_cancelled()

            if response.status_code != 200:
                return False

            data = response.json()
            comments = data.get("comments", [])

            for comment in comments:
                self.check_cancelled()
                author = comment.get("author", {})

                if self._is_current_user(author):
                    created_str = comment.get("created", "")
                    if created_str:
                        comment_dt = self._parse_jira_datetime(created_str)
                        if not comment_dt:
                            continue
                        comment_date = comment_dt.date()
                        if start_date <= comment_date <= end_date:
                            return True

            return False
        except OperationCancelled:
            raise
        except Exception:
            return False

    def user_commented_within_months(self, issue_key, months=3):
        """Check if current user commented on this issue within the last N months"""
        try:
            self.check_cancelled()
            url = f"{self.base_url}/rest/api/2/issue/{issue_key}/comment"
            response = self.session.get(url, timeout=30)
            self.check_cancelled()

            if response.status_code != 200:
                return False

            data = response.json()
            comments = data.get("comments", [])

            since_date = datetime.date.today() - timedelta(days=months * 30)

            for comment in comments:
                self.check_cancelled()
                author = comment.get("author", {})

                if self._is_current_user(author):
                    created_str = comment.get("created", "")
                    if created_str:
                        comment_dt = self._parse_jira_datetime(created_str)
                        if not comment_dt:
                            continue
                        comment_date = comment_dt.date()
                        if comment_date >= since_date:
                            return True

            return False
        except OperationCancelled:
            raise
        except Exception:
            return False

    def get_user_latest_comment(self, issue_key, start_date, end_date):
        try:
            self.check_cancelled()
            url = f"{self.base_url}/rest/api/2/issue/{issue_key}/comment"
            response = self.session.get(url, timeout=30)
            self.check_cancelled()

            if response.status_code != 200:
                return None

            data = response.json()
            comments = data.get("comments", [])

            latest_comment = None
            latest_dt = None

            for comment in comments:
                self.check_cancelled()
                author = comment.get("author", {})

                if self._is_current_user(author):
                    created_str = comment.get("created", "")
                    if created_str:
                        comment_dt = self._parse_jira_datetime(created_str)
                        if not comment_dt:
                            continue
                        comment_date = comment_dt.date()
                        if start_date <= comment_date <= end_date:
                            body = comment.get("body", "")
                            text = self._clean_comment_for_display(body)
                            if text and (latest_dt is None or comment_dt > latest_dt):
                                latest_dt = comment_dt
                                if self.comment_timestamp_prefix_var.get():
                                    latest_comment = f"[{comment_dt.strftime('%m-%d %H:%M')}] {text}"
                                else:
                                    latest_comment = text

            return latest_comment
        except OperationCancelled:
            raise
        except Exception:
            return None

    def get_all_comments_in_range(self, issue_key, start_date, end_date, context_start=None):
        """Get ALL comments (from any author) within the date range for an issue.

        context_start: if provided, also include comments from context_start to start_date-1 as
                       background context (marked with in_period=False).  This allows the AI to
                       understand solutions proposed before the strict report window.
        """
        try:
            self.check_cancelled()
            url = f"{self.base_url}/rest/api/2/issue/{issue_key}/comment"
            response = self.session.get(url, timeout=30)
            self.check_cancelled()

            if response.status_code != 200:
                return []

            data = response.json()
            comments = data.get("comments", [])

            fetch_start = context_start if context_start else start_date

            result = []
            for comment in comments:
                self.check_cancelled()
                author = comment.get("author", {})
                # 优先使用账号名(name，通常是英文)，fallback到displayName
                author_name = author.get("name") or author.get("displayName", "Unknown")
                author_email = author.get("emailAddress", "")
                created_str = comment.get("created", "")
                if not created_str:
                    continue
                comment_dt = self._parse_jira_datetime(created_str)
                if not comment_dt:
                    continue
                comment_date = comment_dt.date()
                if fetch_start <= comment_date <= end_date:
                    body = comment.get("body", "") or ""
                    # Full cleaning via shared method (HTML + noise patterns + whitespace)
                    text = self._clean_comment_body(body)
                    if text:
                        result.append({
                            "author": author_name,
                            "author_role": self._get_comment_author_role(author_name, author_email),
                            "date": comment_date,
                            "body": text,
                            "in_period": start_date <= comment_date <= end_date,
                        })

            return result
        except OperationCancelled:
            raise
        except Exception:
            return []

    def _clean_comment_body(self, body):
        """Clean comment body: remove HTML, markup, attachments, emoji and system-generated noise.
        This is the single source of truth for comment sanitisation — called at data-collection time
        so all downstream consumers (AI prompt, fallback, batch) always receive clean text."""
        if not body:
            return ""

        # --- HTML cleanup ---
        body = re.sub(r'<div[^>]*>.*?</div>', '', body, flags=re.DOTALL)
        body = re.sub(r'<img[^>]*>', '', body)
        body = re.sub(r'<a[^>]*href=[^>]*>[^<]*</a>', '', body)
        body = re.sub(r'<[^>]+>', '', body)
        body = (body.replace('&nbsp;', ' ').replace('&amp;', '&')
                    .replace('&lt;', '<').replace('&gt;', '>')
                    .replace('&quot;', '"').replace('&#39;', "'"))

        # --- Jira wiki markup ---
        # Remove Jira macro blocks: {panel:...}...{panel}, {code}...{code}, {noformat}...{noformat}
        body = re.sub(r'\{(?:panel|code|noformat|color|quote)[^}]*\}.*?\{(?:panel|code|noformat|color|quote)\}', '', body, flags=re.DOTALL | re.IGNORECASE)
        # Remove remaining single-brace macros {macro:...} or {macro}
        body = re.sub(r'\{[a-z][^}]{0,40}\}', '', body, flags=re.IGNORECASE)
        # Remove Jira user mentions [~username]
        body = re.sub(r'\[~[^\]]+\]', '', body)
        # Remove Jira heading markers (keep text): h1. h2. ...
        body = re.sub(r'^h[1-6]\.\s*', '', body, flags=re.MULTILINE)
        # Strip bold/italic/monospace markers (*text*, _text_, +text+, -text-, ^text^, ~text~)
        body = re.sub(r'[*_+\-^~](\S[^*_+\-^~\n]*?\S)[*_+\-^~]', r'\1', body)

        # --- Attachments / images ---
        body = re.sub(r'![\w\-\. ]+(?:\|[^!]*)!', '', body)        # !filename! or !filename|thumbnail!
        body = re.sub(r'https?://\S+?\.(?:jpg|jpeg|png|gif|webp|bmp|svg)(?:\?\S*)?', '', body, flags=re.IGNORECASE)
        body = re.sub(r'\[\^[^\]]+\]', '', body)                   # [^attachment.zip]
        body = re.sub(r'_\(\d+[KMG]?B?\)_', '', body)             # _(32KB)_
        body = re.sub(r'\[[^\]]*\.(?:zip|gz|rar|7z|tar|dmp|log|txt|pdf|doc|docx|xls|xlsx|ild)[^\]]*\]', '', body, flags=re.IGNORECASE)
        body = re.sub(r'(?:attachment|附件|文件)[：:\s][^\s,，\n]+', '', body, flags=re.IGNORECASE)

        # --- Emoji (unicode ranges) ---
        body = re.compile(
            "["
            "\U0001F600-\U0001F64F\U0001F300-\U0001F5FF\U0001F680-\U0001F6FF"
            "\U0001F700-\U0001F77F\U0001F780-\U0001F7FF\U0001F800-\U0001F8FF"
            "\U0001F900-\U0001F9FF\U0001FA00-\U0001FA6F\U0001FA70-\U0001FAFF"
            "\u2600-\u27BF"
            "]+", flags=re.UNICODE
        ).sub('', body)

        # --- System-generated / noise phrases (entire sentence removed) ---
        _NOISE_PATTERNS = [
            # Parent-issue sync creation — Chinese variants (covers "通过复制")
            r'(?:确认)?该issue(?:已)?由父issue\s*\S+\s*(?:通过复制\s*)?同步创建[^\n。]*',
            r'从父(?:issue|任务)\s*(?:\S+\s*)?(?:通过复制\s*)?同步创建[^\n。]*',
            r'由父issue[^\n。]*同步创建[^\n。]*',
            r'通过复制同步创建[^\n。]*',
            # Parent-issue sync creation — English variants
            r'[Tt]his issue was created from parent issue\s*\S+\s*by a Copy\s*[&＆]?\s*Sync operation[^\n]*',
            r'[Cc]reated from parent issue[^\n]*(?:Copy|Sync)[^\n]*',
            r'[Ss]ynced? (?:from|with) parent issue[^\n]*',
            r'[Tt]his issue requires your attention[^\n]*',
            r'\b(?:Log|dbg|debug|dump|trace)[^。\n]*\.(?:zip|rar|7z|tar|gz|log|dmp|txt)\s*\([^)]*\)',
            # Generic low-value actions — Chinese
            r'报告了一个bug[^\n。]*',
            r'创建了(?:case|issue|工单)[^\n。]*',
            r'在Jira中进行了记录[^\n。]*',
            r'提供了日志用于分析[^\n。]*',
            r'(?:已)?进行了记录[^\n。]*',
            r'确认该issue[^\n。]*',
            r'请(?:帮忙)?关闭(?:该)?(?:case|issue|工单)[^\n。]*',
            r'已回复[^\n。]{0,20}',
            # Jira UI noise (English)
            r'(?:Permalink|Edit|Delete|added a comment)[^\n]*',
        ]
        for pattern in _NOISE_PATTERNS:
            body = re.sub(pattern, '', body, flags=re.IGNORECASE)

        # --- Final whitespace normalisation ---
        body = re.sub(r'\s+', ' ', body).strip()

        # Discard if only punctuation / symbols remain (no CJK or Latin word chars)
        if body and not re.search(r'[\w\u4e00-\u9fff]', body):
            return ""

        return body

    def _get_comment_author_role(self, author_name, author_email):
        """Classify comment authors so the model can distinguish our action vs customer feedback."""
        author_email = (author_email or "").lower()
        author_name = (author_name or "").lower()
        username = (self.username or "").lower()
        username_short = username.split("@")[0] if username else ""

        if username and (author_email == username or author_name == username or author_name == username_short):
            return "当前用户"
        if author_email.endswith("@quectel.com"):
            return "我方"
        return "客户/Reporter"

    def _compact_comment_signal(self, body):
        """Reduce noisy technical text before sending it to the model.

        Keep action/result words, but collapse long paths and attachment-like fragments so the
        model does not summarize a bare path, file name, CFUN/NV token, or log name as progress.
        """
        if not body:
            return ""

        text = body
        text = re.sub(r'(?:[A-Za-z0-9_.-]+[\\/]){2,}[A-Za-z0-9_.-]+', '[路径]', text)
        text = re.sub(r'(?<![A-Za-z0-9_.-])(?:[A-Za-z]:)?(?:[\\/][A-Za-z0-9_. \-]+){2,}', '[路径]', text)
        text = re.sub(r'\b[\w.-]+\.(?:zip|rar|7z|tar|gz|log|dmp|txt)\b\s*(?:\([^)]*\))?', '', text, flags=re.IGNORECASE)
        text = re.sub(r'\s+', ' ', text).strip()

        # Common short-form solution comments: "qlrild 替换[路径]下同文件试下"
        text = re.sub(r'\b([\w.-]{2,})\s+替换\[路径\]下同文件试下', r'提供\1文件替换方案', text, flags=re.IGNORECASE)
        text = re.sub(r'替换\[路径\]下同文件试下', '替换对应目录下同名文件', text)
        text = text.replace('[路径]下同文件', '对应目录下同名文件')
        text = text.replace('[路径]', '对应路径')
        return text.strip()

    def _comment_signal_score(self, comment):
        """Rank comments by progress value; used when prompt length forces truncation."""
        body = comment.get('body', '')
        role = comment.get('author_role', '')
        score = 0
        if comment.get('in_period', True):
            score += 2
        if role in ("当前用户", "我方"):
            score += 2
        if re.search(r'验证可以|验证通过|测试通过|恢复正常|解决|关闭|closed|验证完成|没有问题', body, re.IGNORECASE):
            score += 6
        if re.search(r'提供|替换|修改|配置|方案|补丁|patch|disable|disabled|烧写|排查|确认|说明|建议', body, re.IGNORECASE):
            score += 4
        if re.search(r'\b(?:Log|dbg|dump|trace)\b|日志|附件', body, re.IGNORECASE):
            score -= 2
        return score

    def _normalize_progress_text(self, text):
        """Make fallback/postprocessed summaries read like progress, not copied raw comments."""
        text = self._compact_comment_signal(text)
        text = re.sub(r'^提供([\w.-]{2,})文件替换方案$', r'提供\1文件替换方案', text)
        text = text.replace('验证可以', '验证通过')
        text = text.replace('验证完成，没有问题', '验证无问题')
        text = text.replace('此单关闭', '问题关闭')
        text = re.sub(r'\s+', ' ', text).strip(' ，,。')
        return text

    def _has_resolution_signal(self, text):
        return bool(re.search(r'验证可以|验证通过|测试通过|恢复正常|问题关闭|此单关闭|解决|closed|验证完成|没有问题', text, re.IGNORECASE))

    def _has_solution_signal(self, text):
        return bool(re.search(r'提供|替换|修改|配置|方案|补丁|patch|disable|disabled|烧写|排查|确认|说明|建议|NV文件', text, re.IGNORECASE))

    def _is_low_quality_summary(self, summary):
        """Detect model outputs that are just a token/path/keyword instead of a progress sentence."""
        if not summary:
            return True
        text = summary.strip()
        if "\\" in text or "/" in text:
            return True
        if len(text) <= 18 and re.search(r'[A-Za-z0-9]', text):
            return True
        if re.fullmatch(r'[\w.\-]+', text):
            return True
        return False

    def _sanitize_ai_summary(self, summary, comments):
        """Replace weak AI output with deterministic comment-derived fallback."""
        summary = (summary or "").strip()
        fallback = self._fallback_summary(comments)
        if self._is_low_quality_summary(summary):
            return fallback
        if re.search(r'无进展|仍在排查|无实质进展', summary) and fallback not in ("无评论", "仍在排查中"):
            if any(self._has_solution_signal(c.get('body', '')) or self._has_resolution_signal(c.get('body', '')) for c in comments):
                return fallback
        return summary

    def _format_comments_for_ai(self, comments, max_chars=1500):
        """Format pre-cleaned comments into a compact string for the AI prompt.

        Comments with in_period=True (within the strict report window) are marked [本期].
        Older background-context comments are marked [背景].
        Both are included so the AI has full context while knowing what is recent.
        """
        if not comments:
            return ""

        formatted = []
        total_chars = 0

        ordered_comments = sorted(
            comments,
            key=lambda c: (self._comment_signal_score(c), c.get('date')),
            reverse=True
        )

        for c in ordered_comments:
            body = self._compact_comment_signal(c['body'])
            if not body:
                continue

            date_str = c['date'].strftime("%m-%d") if hasattr(c['date'], 'strftime') else str(c['date'])
            tag = "[本期]" if c.get('in_period', True) else "[背景]"
            role = c.get('author_role', '评论')

            # Keep each comment concise; 250 chars is enough for one update
            if len(body) > 250:
                body = body[:250] + "…"

            entry = f"{tag}[{role}/{c['author']} {date_str}] {body}"
            if total_chars + len(entry) > max_chars:
                remaining = max_chars - total_chars
                if remaining > 80:
                    formatted.append(entry[:remaining] + "…")
                break

            formatted.append(entry)
            total_chars += len(entry) + 1

        return "\n".join(formatted)

    def summarize_progress_with_ai(self, issue_key, summary, comments, model):
        """Use DeepSeek AI to summarize issue progress from comments"""
        self.check_cancelled()
        api_key = self.saved_deepseek_api_key.strip()
        if not api_key:
            return "[AI总结] 未配置API Key"

        if not comments:
            return "[AI总结] 无评论"

        # Use improved formatting method
        comments_text = self._format_comments_for_ai(comments, max_chars=900)

        has_in_period = any(c.get('in_period', True) for c in comments)
        period_hint = (
            "[本期]=本报告周期内评论，[背景]=周期前背景。优先基于[本期]总结，无[本期]时用[背景]。"
            if any(not c.get('in_period', True) for c in comments) else ""
        )

        prompt = (
            f"Issue: {summary}\n"
            f"{'标注说明: ' + period_hint if period_hint else ''}\n\n"
            f"{comments_text}\n\n"
            f"用1~3句话总结技术进展，优先级：验证/恢复/关闭结果 > 当前用户或我方提供的方案/文件/补丁 > 分析结论 > 待确认。"
            f"必须写成“动作+结果/状态”，不要输出裸路径、文件名、NV/CFUN关键词或日志名。"
            f"若客户回复验证可以/恢复正常，要明确写验证通过/问题关闭；若无实质进展才回复【仍在排查中】。"
        ).strip()

        try:
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            payload = {
                "model": model,
                "messages": [
                    {"role": "user", "content": prompt}
                ],
                "max_tokens": 500,
                "temperature": 0.3
            }
            response = requests.post(
                "https://api.deepseek.com/chat/completions",
                headers=headers,
                json=payload,
                timeout=60
            )
            self.check_cancelled()

            # Try to parse JSON first
            try:
                result = response.json()
            except ValueError:
                # Response is not JSON, check status
                if response.status_code == 200:
                    # Empty or malformed response
                    return self._fallback_summary(comments, "API返回格式错误")
                return self._fallback_summary(comments, f"API返回非JSON: HTTP {response.status_code}")

            # Check for API-level errors in response
            if "error" in result:
                return self._fallback_summary(comments, "API返回错误")

            if response.status_code == 200:
                choices = result.get("choices")
                if not choices:
                    return self._fallback_summary(comments, "API返回空")

                message = choices[0].get("message", {})
                summary_text = message.get("content", "").strip()

                if summary_text:
                    return self._sanitize_ai_summary(summary_text, comments)
                else:
                    return self._fallback_summary(comments)
            elif response.status_code == 401:
                return "[AI总结] DeepSeek API Key无效"
            elif response.status_code == 429:
                return self._fallback_summary(comments, "API请求超过限额")
            elif response.status_code == 400:
                return self._fallback_summary(comments, "请求参数错误")
            else:
                return self._fallback_summary(comments, f"API错误: HTTP {response.status_code}")

        except requests.exceptions.Timeout:
            return self._fallback_summary(comments)
        except requests.exceptions.ConnectionError:
            return self._fallback_summary(comments)
        except OperationCancelled:
            raise
        except Exception:
            return self._fallback_summary(comments)

        return self._fallback_summary(comments)

    def _fallback_summary(self, comments, reason=""):
        """Deterministic fallback summary that prefers solution + verification/resolution."""
        if not comments:
            return "无评论"

        sorted_comments = sorted(comments, key=lambda x: x['date'])
        outcome = None
        solution = None
        latest_signal = None

        for comment in sorted_comments:
            body = self._normalize_progress_text(comment.get('body', ''))
            if not body:
                continue
            if self._has_solution_signal(body):
                solution = body
                latest_signal = body
            if self._has_resolution_signal(body):
                outcome = body
                latest_signal = body

        if outcome:
            if len(outcome) <= 8 and solution:
                return f"{solution}，{outcome}。"
            return outcome[:120] + ("..." if len(outcome) > 120 else "")

        if solution:
            return solution[:120] + ("..." if len(solution) > 120 else "")

        if latest_signal:
            return latest_signal[:120] + ("..." if len(latest_signal) > 120 else "")

        # Last resort: latest cleaned comment
        latest = sorted(comments, key=lambda x: x['date'], reverse=True)[0]
        body = self._normalize_progress_text(latest.get('body', ''))

        # Truncate if too long
        if len(body) > 100:
            body = body[:100] + "..."

        return body

    def _load_skill(self, skill_name):
        """Load a skill markdown file and return its content."""
        skill_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "skills", f"{skill_name}.md")
        if os.path.exists(skill_path):
            try:
                with open(skill_path, "r", encoding="utf-8") as f:
                    return f.read()
            except:
                pass
        return None

    def batch_summarize_with_ai(self, issues_data, model):
        """Batch summarize multiple issues with a single API call

        issues_data: list of dicts with {issue_key, summary, comments}
        Returns: dict mapping issue_key to summary
        """
        self.check_cancelled()
        api_key = self.saved_deepseek_api_key.strip()
        if not api_key:
            return {item['issue_key']: "未配置API Key" for item in issues_data}

        # Filter items with comments
        items_with_comments = [item for item in issues_data if item['comments']]
        if not items_with_comments:
            return {item['issue_key']: "无评论" for item in issues_data}

        def batch_fallback():
            return {
                item['issue_key']: self._fallback_summary(item['comments']) if item['comments'] else "无评论"
                for item in issues_data
            }

        # Build combined prompt
        combined_text = []
        has_context_tags = False
        for item in items_with_comments:
            comments_text = self._format_comments_for_ai(item['comments'], max_chars=600)
            title_line = f"【{item['summary']}】" if item.get('summary') else ""
            combined_text.append(f"## {item['issue_key']} {title_line}\n{comments_text}")
            if any(not c.get('in_period', True) for c in item['comments']):
                has_context_tags = True

        period_note = "[本期]=报告周期内, [背景]=周期前背景。优先基于[本期]总结。\n" if has_context_tags else ""
        issues_block = "\n\n".join(combined_text)

        # Load skill for prompt template
        skill_content = self._load_skill("batch_ai_summary_skill")
        if skill_content:
            prompt = f"{period_note}{skill_content}\n\n{issues_block}"
        else:
            prompt = (
                f"{period_note}"
                f"总结以下每个Jira issue技术进展，每项1~2句话。优先级：验证/恢复/关闭结果 > 当前用户或我方方案/文件/补丁 > 分析结论 > 待确认。\n"
                f"必须写成动作+结果/状态，不要输出裸路径、文件名、NV/CFUN关键词或日志名。\n"
                f"格式：issue_key: 总结内容（无实质进展才写：issue_key: 仍在排查中）\n\n"
                f"{issues_block}"
            )

        try:
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            payload = {
                "model": model,
                "messages": [
                    {"role": "user", "content": prompt}
                ],
                "max_tokens": 2000,
                "temperature": 0.3
            }
            response = requests.post(
                "https://api.deepseek.com/chat/completions",
                headers=headers,
                json=payload,
                timeout=180
            )
            self.check_cancelled()

            try:
                result = response.json()
            except ValueError:
                return batch_fallback()

            if "error" in result:
                return batch_fallback()

            if response.status_code == 200:
                choices = result.get("choices", [])
                if choices:
                    content = choices[0].get("message", {}).get("content", "").strip()
                    return self._parse_batch_results(content, items_with_comments)

        except OperationCancelled:
            raise
        except Exception:
            return batch_fallback()

        return batch_fallback()

    def _parse_batch_results(self, content, items):
        """Parse batch AI response into individual summaries"""
        results = {}
        lines = content.split('\n')

        for line in lines:
            line = line.strip()
            if not line:
                continue
            line = re.sub(r'^[-*•\d.\s]+', '', line).strip()
            # Match "issue_key: summary" format, accepting full-width colon and markdown noise
            match = re.match(r'([A-Z]+-\d+)\s*[:：]\s*(.+)$', line)
            if match:
                key = match.group(1).strip()
                summary = match.group(2).strip()
                item = next((item for item in items if item['issue_key'] == key), None)
                results[key] = self._sanitize_ai_summary(summary, item['comments']) if item else summary

        # Check if we got any valid results
        for item in items:
            key = item['issue_key']
            if key not in results:
                # Fallback to latest comment
                comments = item['comments']
                if comments:
                    results[key] = self._fallback_summary(comments)
                else:
                    results[key] = "无评论"

        return results

    def create_excel(self, issues, filepath, statuses, start_date, end_date):
        self.check_cancelled()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        font_chinese = Font(name="Microsoft YaHei", size=10)
        font_english = Font(name="JetBrains Mono", size=10)
        font_header = Font(name="Microsoft YaHei", bold=True, color="FF000000", size=10)

        header_align = self.header_align_var.get()
        cell_align = self.cell_align_var.get()
        header_alignment = Alignment(horizontal=header_align, vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal=cell_align, vertical="center", wrap_text=True)

        thin = Side(border_style="thin", color="FF000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        col_order = [int(x.strip()) - 1 for x in self.column_order_var.get().split(",")]

        header_names = ["客户名称", "型号", "问题描述",
                       "JIRA号", "状态", "是否为重点问题",
                       "进展"]

        def has_chinese(text):
            return any('一' <= c <= '鿿' for c in str(text))

        def set_cell_font(cell, value):
            text = str(value) if value is not None else ""
            cell.value = text if text else value
            cell.font = font_chinese if has_chinese(text) else font_english

        for col, idx in enumerate(col_order, 1):
            header = header_names[idx]
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = font_header
            cell.alignment = header_alignment
            cell.border = border

        status_col = col_order.index(4) + 1
        key_issue_col = col_order.index(5) + 1
        status_range = f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{len(issues) + 1}"
        key_issue_range = f"{get_column_letter(key_issue_col)}2:{get_column_letter(key_issue_col)}{len(issues) + 1}"

        ws_options = wb.create_sheet("_Options")
        ws_options.sheet_state = "hidden"
        for i, opt in enumerate(["WAIT FAE INFO", "WORKED AROUND", "WORKING", "CLOSED", "RESOLVED", "WAIT 3RD PARTY"], 1):
            ws_options.cell(row=i, column=1, value=opt)
        for i, opt in enumerate(["是", "否"], 1):
            ws_options.cell(row=i, column=2, value=opt)
        status_options_range = f"_Options!$A$1:$A$6"
        key_issue_options_range = f"_Options!$B$1:$B$2"

        dv_status = DataValidation(type="list", formula1=status_options_range, allow_blank=True)
        dv_status.error = "Please select a valid status"
        dv_status.errorTitle = "Invalid Status"
        ws.add_data_validation(dv_status)
        dv_status.sqref = status_range

        dv_key_issue = DataValidation(type="list", formula1=key_issue_options_range, allow_blank=True)
        dv_key_issue.error = "Please select 是 or 否"
        dv_key_issue.errorTitle = "Invalid Key Issue"
        ws.add_data_validation(dv_key_issue)
        dv_key_issue.sqref = key_issue_range

        # Pre-fetch all comments if AI summary is enabled
        latest_comments = {}
        if self.use_ai_summary_var.get():
            self.root.after(0, lambda: self.update_processing("Fetching comments...", "", 72))
            # Use a 60-day context window so solutions given before the report period are visible
            context_start = end_date - timedelta(days=60)
            issues_data = []
            total_issues = max(len(issues), 1)
            for idx, issue in enumerate(issues, 1):
                self.check_cancelled()
                issue_key = issue.get("key", "")
                fields = issue.get("fields", {})
                progress = 72 + (idx / total_issues) * 8
                self.root.after(0, lambda k=issue_key, p=progress: self.update_processing("Fetching comments...", k, p))
                all_comments = self.get_all_comments_in_range(
                    issue_key, start_date, end_date, context_start=context_start
                )
                issues_data.append({
                    "issue_key": issue_key,
                    "summary": fields.get("summary", ""),
                    "comments": all_comments
                })

            if self.batch_mode_var.get():
                # Batch mode: process all at once
                batch_size = max(1, self.batch_size_var.get())
                self.root.after(0, lambda: self.update_processing(f"AI batch summarizing ({len(issues_data)} issues)...", "", 80))
                total_batches = max((len(issues_data) + batch_size - 1) // batch_size, 1)
                for batch_index, i in enumerate(range(0, len(issues_data), batch_size), 1):
                    self.check_cancelled()
                    batch = issues_data[i:i+batch_size]
                    progress = 80 + (batch_index / total_batches) * 10
                    self.root.after(0, lambda b=batch_index, t=total_batches, p=progress: self.update_processing("AI batch summarizing...", f"Batch {b}/{t}", p))
                    results = self.batch_summarize_with_ai(batch, self.ai_model_var.get())
                    latest_comments.update(results)
            else:
                # Individual mode: process one by one
                total_ai = max(len(issues_data), 1)
                for idx, item in enumerate(issues_data, 1):
                    self.check_cancelled()
                    progress = 80 + (idx / total_ai) * 10
                    self.root.after(0, lambda k=item['issue_key'], i=idx, t=total_ai, p=progress: self.update_processing(f"AI summarizing {k}...", f"{i}/{t}", p))
                    ai_summary = self.summarize_progress_with_ai(
                        item['issue_key'],
                        item['summary'],
                        item['comments'],
                        self.ai_model_var.get()
                    )
                    latest_comments[item['issue_key']] = ai_summary

        total_rows = max(len(issues), 1)
        for row, issue in enumerate(issues, 2):
            self.check_cancelled()
            fields = issue.get("fields", {})
            issue_key = issue.get("key", "")
            customer_name, model_name = self._resolve_customer_and_model(issue_key, fields)

            latest_comment = latest_comments.get(issue_key, "")
            if self.fetch_comment_var.get() and not latest_comment:
                progress = 80 + ((row - 1) / total_rows) * 10
                self.root.after(0, lambda k=issue_key, p=progress: self.update_processing(f"Fetching comment for {k}...", "", p))
                latest_comment = self.get_user_latest_comment(issue_key, start_date, end_date) or ""

            values = [
                customer_name,
                model_name,
                fields.get("summary", ""),
                issue_key,
                fields.get("status", {}),
                fields.get("priority", {}),
                latest_comment,
            ]

            for col, idx in enumerate(col_order, 1):
                val = values[idx]
                if idx == 4:
                    val = val.get("name", "") if isinstance(val, dict) else val
                elif idx == 5:
                    priority_name = val.get("name", "") if isinstance(val, dict) else val
                    val = "是" if priority_name in ("Highest", "High") else "否"

                cell = ws.cell(row=row, column=col, value=val)
                if idx == 3:
                    cell.hyperlink = f"{self.base_url}/browse/{issue_key}"
                set_cell_font(cell, val)
                cell.alignment = cell_alignment
                cell.border = border
                if idx == 5 and val == "是" and self.key_issue_highlight_var.get():
                    cell.fill = PatternFill(fill_type="solid", start_color="FFFFC7CE", end_color="FFFFC7CE")

            progress = 90 + ((row - 1) / total_rows) * 8
            self.root.after(0, lambda k=issue_key, p=progress: self.update_processing("Writing Excel rows...", k, p))

        for col in range(1, 8):
            self.check_cancelled()
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        max_length = max(max_length, cell_len)
                except:
                    pass
            adjusted_width = min(max_length + 5, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.row_dimensions[1].height = 25

        self.check_cancelled()
        self.root.after(0, lambda: self.update_processing("Saving Excel file...", os.path.basename(filepath), 99))
        wb.save(filepath)


def main():
    root = tk.Tk()
    app = JiraReportApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
